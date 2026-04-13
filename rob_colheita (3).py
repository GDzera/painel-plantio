"""
╔══════════════════════════════════════════════════════════════╗
║  ROB_COLHEITA v2 — PAINEL INTEGRADO CCT IPAUSSU             ║
║  Produção + Perdas + Overview CCT (Métricas de Transporte)   ║
╚══════════════════════════════════════════════════════════════╝
"""

import pandas as pd
import json, sys, os, time, numpy as np, traceback
from pathlib import Path
from datetime import datetime
from collections import defaultdict
# ═══════════════════════════════════════════════════════════
#  SISTEMA DE AUTENTICAÇÃO E CRIPTOGRAFIA
# ═══════════════════════════════════════════════════════════
import hashlib, base64, argparse
USUARIOS_FILE = "usuarios.json"

def _evp_bytes_to_key(password, salt, key_len=32, iv_len=16):
    """Deriva key+IV compatível com CryptoJS (EVP_BytesToKey/MD5)"""
    dtot, d = b'', b''
    while len(dtot) < key_len + iv_len:
        d = hashlib.md5(d + password + salt).digest()
        dtot += d
    return dtot[:key_len], dtot[key_len:key_len+iv_len]

def aes_encrypt(plaintext, passphrase):
    """AES-256-CBC compatível com CryptoJS.AES.encrypt(data, passphrase)"""
    from Crypto.Cipher import AES as _AES
    salt = os.urandom(8)
    key, iv = _evp_bytes_to_key(passphrase.encode(), salt)
    raw = plaintext.encode('utf-8')
    pad_len = 16 - (len(raw) % 16)
    padded = raw + bytes([pad_len] * pad_len)
    cipher = _AES.new(key, _AES.MODE_CBC, iv)
    ct = cipher.encrypt(padded)
    return base64.b64encode(b'Salted__' + salt + ct).decode()

def load_usuarios():
    if os.path.exists(USUARIOS_FILE):
        with open(USUARIOS_FILE, 'r') as f:
            return json.load(f)
    return {"master_key": base64.b64encode(os.urandom(32)).decode(), "users": {}}

def save_usuarios(data):
    with open(USUARIOS_FILE, 'w') as f:
        json.dump(data, f, indent=2)

def add_usuario(username, password):
    db = load_usuarios()
    mk = db["master_key"]
    pw_hash = hashlib.sha256(password.encode()).hexdigest()
    mk_encrypted = aes_encrypt(mk, password)
    db["users"][username] = {"hash": pw_hash, "mk_enc": mk_encrypted}
    save_usuarios(db)
    print(f"  ✅ Usuário \'{username}\' adicionado com sucesso.")

def del_usuario(username):
    db = load_usuarios()
    if username in db["users"]:
        del db["users"][username]
        save_usuarios(db)
        print(f"  ✅ Usuário \'{username}\' removido.")
    else:
        print(f"  ❌ Usuário \'{username}\' não encontrado.")

def list_usuarios():
    db = load_usuarios()
    if not db["users"]:
        print("  Nenhum usuário cadastrado. Use: python rob_colheita.py --add usuario senha")
        return
    print(f"  {'Usuário':<20} {'Status'}")
    print(f"  {'-'*20} {'-'*10}")
    for u in db["users"]:
        print(f"  {u:<20} ativo")
    print(f"\n  Total: {len(db['users'])} usuário(s)")

def encrypt_dashboard(html):
    """Criptografa todo o conteúdo do dashboard e insere tela de login"""
    db = load_usuarios()
    if not db["users"]:
        print("  ⚠️  Nenhum usuário cadastrado! Dashboard será gerado SEM proteção.")
        print("     Use: python rob_colheita.py --add admin suasenha")
        return html

    mk = db["master_key"]
    users_js = {}
    for u, info in db["users"].items():
        users_js[u] = {"h": info["hash"], "mk": info["mk_enc"]}

    # Extrair e criptografar o body
    body_start = html.index('<body>') + 6
    body_end = html.index('</body>')
    body_content = html[body_start:body_end]
    encrypted_body = aes_encrypt(body_content, mk)

    login_page = _build_login_page(encrypted_body, users_js)
    return html[:body_start] + login_page + html[body_end:]

def _build_login_page(encrypted_data, users_js):
    users_json = json.dumps(users_js)
    return """
<style>
.login-overlay{position:fixed;top:0;left:0;width:100%;height:100%;background:linear-gradient(135deg,#004D23 0%,#00843D 50%,#006830 100%);display:flex;align-items:center;justify-content:center;z-index:9999;font-family:'Plus Jakarta Sans',sans-serif}
.login-box{background:#fff;border-radius:20px;padding:48px 40px;width:380px;box-shadow:0 20px 60px rgba(0,0,0,.3);text-align:center}
.login-logo{width:64px;height:64px;background:#00843D;border-radius:16px;display:inline-flex;align-items:center;justify-content:center;font-size:28px;font-weight:900;color:#fff;margin-bottom:20px;font-family:'JetBrains Mono',monospace}
.login-title{font-size:22px;font-weight:800;color:#1a2e22;margin-bottom:6px}
.login-sub{font-size:13px;color:#5a7265;margin-bottom:28px}
.login-input{width:100%;padding:14px 16px;border:2px solid #e0e8e3;border-radius:10px;font-size:14px;font-family:'Plus Jakarta Sans',sans-serif;margin-bottom:12px;outline:none;transition:border-color .2s;box-sizing:border-box}
.login-input:focus{border-color:#00843D}
.login-btn{width:100%;padding:14px;border:none;border-radius:10px;background:#00843D;color:#fff;font-size:15px;font-weight:700;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;transition:background .2s;margin-top:4px}
.login-btn:hover{background:#006830}
.login-error{color:#d63031;font-size:12px;font-weight:600;margin-top:12px;min-height:18px}
.login-footer{font-size:10px;color:rgba(255,255,255,.5);position:fixed;bottom:16px;left:0;right:0;text-align:center}
.login-lock{font-size:11px;color:#5a7265;margin-top:16px;display:flex;align-items:center;justify-content:center;gap:4px}
</style>
<script src="https://cdnjs.cloudflare.com/ajax/libs/crypto-js/4.2.0/crypto-js.min.js"></script>

<div class="login-overlay" id="loginOverlay">
  <div class="login-box">
    <div class="login-logo">R</div>
    <div class="login-title">Gerencial Ipaussu</div>
    <div class="login-sub">Painel Integrado — Acesso Restrito</div>
    <input class="login-input" id="loginUser" type="text" placeholder="Usuário" autocomplete="username" onkeydown="if(event.key==='Enter')document.getElementById('loginPass').focus()">
    <input class="login-input" id="loginPass" type="password" placeholder="Senha" autocomplete="current-password" onkeydown="if(event.key==='Enter')doLogin()">
    <button class="login-btn" onclick="doLogin()">Entrar</button>
    <div class="login-error" id="loginError"></div>
    <div class="login-lock">🔒 Dados criptografados com AES-256</div>
  </div>
  <div class="login-footer">Ambiente protegido • Dados criptografados end-to-end</div>
</div>

<div id="dashContent" style="display:none"></div>

<script>
const _U=""" + users_json + """;
const _E='""" + encrypted_data + """';

function sha256(str){return CryptoJS.SHA256(str).toString()}

function doLogin(){
  const user=document.getElementById('loginUser').value.trim().toLowerCase();
  const pass=document.getElementById('loginPass').value;
  const errEl=document.getElementById('loginError');

  if(!user||!pass){errEl.textContent='Preencha usuário e senha.';return}

  const uData=_U[user];
  if(!uData){errEl.textContent='Usuário não encontrado.';return}

  const passHash=sha256(pass);
  if(passHash!==uData.h){errEl.textContent='Senha incorreta.';return}

  try{
    // Decrypt master key with user's password
    const mkDecrypted=CryptoJS.AES.decrypt(uData.mk,pass).toString(CryptoJS.enc.Utf8);
    if(!mkDecrypted){errEl.textContent='Erro na descriptografia.';return}

    // Decrypt dashboard body with master key
    const bodyDecrypted=CryptoJS.AES.decrypt(_E,mkDecrypted).toString(CryptoJS.enc.Utf8);
    if(!bodyDecrypted){errEl.textContent='Erro ao carregar dados.';return}

    // Inject and show
    document.getElementById('loginOverlay').style.display='none';
    document.getElementById('dashContent').style.display='block';
    document.getElementById('dashContent').innerHTML=bodyDecrypted;

    // Execute ONLY executable scripts (skip JSON data containers)
    document.getElementById('dashContent').querySelectorAll('script').forEach(old=>{
      if(old.type && old.type !== 'text/javascript' && old.type !== '') return;
      const s=document.createElement('script');
      if(old.src){s.src=old.src}else{s.textContent=old.textContent}
      old.parentNode.replaceChild(s,old);
    });
    // Re-fire init for scripts that use DOMContentLoaded
    window.dispatchEvent(new Event('DOMContentLoaded'));

    // Session marker
    sessionStorage.setItem('cct_auth','1');
  }catch(e){
    console.error(e);
    errEl.textContent='Falha na autenticação. Tente novamente.';
  }
}

// Auto-focus
document.getElementById('loginUser').focus();
</script>
"""



try:
    from openpyxl import load_workbook
except ImportError:
    print("Instale: pip install openpyxl pandas")
    sys.exit(1)

# ═══════════════════════════════════════════════════════════
#  CONFIGURAÇÃO
# ═══════════════════════════════════════════════════════════
ARQUIVO_HTML     = "dash_colheita.html"
ARQUIVO_PRODUCAO = "CCT_Safra_25_26_Dashboard_Raizen.xlsx"
ARQUIVO_METRICAS = "Metricas_cct_dashboard.xlsx"
INTERVALO_SEG    = 5
ABRIR_BROWSER    = True

META_PERDAS = 3.50
TOLERANCIA_PERDAS = 0.35
COLUNAS_PERDAS = ['AGR_TOUCEIRA','AGR_PALMITO','CANA_INTEIRA','REBOLO','PEDACO','ESTILHACO','LASCA']
TURNOS_VALIDOS = ['A','B','C']
MESES = ['Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
FRENTE_COLORS = [
    ('#00843D','rgba(0,132,61,.15)'), ('#e17055','rgba(225,112,85,.15)'),
    ('#0984e3','rgba(9,132,227,.15)'), ('#6c5ce7','rgba(108,92,231,.15)'),
    ('#00b894','rgba(0,184,148,.15)'), ('#fdcb6e','rgba(253,203,110,.15)'),
]

def sr(v, d=2):
    if v is None: return None
    try: return round(float(v), d)
    except: return None

def sf(v, d=2):
    if v is None: return 0
    try: return round(float(v), d)
    except: return 0

def extrair_frente_valida(val):
    if pd.isnull(val): return None
    s = str(val).strip().replace(".0","")
    if s.isdigit() and int(s) > 0: return s
    return None

# ============================================================
# EXTRATOR 1: PRODUÇÃO
# ============================================================
def extract_producao(xlsx_path):
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    if "Dados diários" not in wb.sheetnames:
        wb.close(); return "{}", []
    ws = wb["Dados diários"]
    row6 = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]
    row7 = list(ws.iter_rows(min_row=7, max_row=7, values_only=True))[0]
    frente_cols = {}
    for i, val in enumerate(row6):
        fn = extrair_frente_valida(val)
        if fn:
            if fn not in frente_cols: frente_cols[fn] = []
            frente_cols[fn].append(i)
    frentes = []
    for fn, cols in frente_cols.items():
        info = {'num':fn,'real':None,'atr':None,'imp':None,'raio':None,'dc':None}
        for c in cols:
            h = str(row7[c]).strip().lower().replace('\n',' ') if c < len(row7) and row7[c] else ''
            if h=='real': info['real']=c
            elif h=='atr': info['atr']=c
            elif 'impureza' in h: info['imp']=c
            elif 'raio' in h: info['raio']=c
            elif 'dens' in h: info['dc']=c
        if info['real'] is not None: frentes.append(info)
    frentes.sort(key=lambda x: x['num'])
    frente_names = [f['num'] for f in frentes]
    rows, metas = [], {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=260, values_only=True)):
        v = list(row)
        if i==1: metas['meta_ipa']={m:sr(v[3+j]) for j,m in enumerate(MESES)}; metas['meta_ipa']['Total']=sr(v[12])
        elif i==2: metas['real_u']={m:sr(v[3+j]) for j,m in enumerate(MESES)}; metas['real_u']['Total']=sr(v[12])
        elif i==3: metas['meta_p']={m:sr(v[3+j]) for j,m in enumerate(MESES)}; metas['meta_p']['Total']=sr(v[12])
        elif i==4: metas['real_p']={m:sr(v[3+j]) for j,m in enumerate(MESES)}; metas['real_p']['Total']=sr(v[12])
        elif i>=7:
            mes,dia,mt = v[0],v[1],v[2]
            if not mes or not dia or mt is None or not isinstance(mt,(int,float)): continue
            e = {'mes':str(mes),'dia':str(dia)[:10],'mt':sr(mt),'mp':sr(v[3]),
                 'mix_p':sr(v[4]*100,1) if isinstance(v[4],(int,float)) else None}
            for f in frentes:
                fn=f['num']
                if f['real'] is not None: e[f'r{fn}']=sr(v[f['real']])
                if f['atr'] is not None: e[f'a{fn}']=sr(v[f['atr']])
                if f['imp'] is not None: e[f'i{fn}']=sr(v[f['imp']])
                if f['raio'] is not None: e[f'k{fn}']=sr(v[f['raio']])
                if f['dc'] is not None: e[f'd{fn}']=sr(v[f['dc']])
            rows.append(e)
    wb.close()
    clean = [{k:val for k,val in d.items() if val is not None} for d in rows]
    return json.dumps({'d':clean,'m':metas}, separators=(',',':')), frente_names

# ============================================================
# EXTRATOR 2: PERDAS
# ============================================================
def extract_perdas(caminho_excel):
    try:
        xl = pd.ExcelFile(caminho_excel)
        if "Base colheita" not in xl.sheet_names: return "null"
        df = pd.read_excel(caminho_excel, sheet_name="Base colheita")
        c_equip = next((c for c in df.columns if 'equip' in str(c).lower()), 'EQUIP')
        c_frente = next((c for c in df.columns if 'frente' in str(c).lower()), 'FRENTE')
        c_turno = next((c for c in df.columns if 'turno' in str(c).lower()), 'TURNO')
        col_perdas_reais = [c for c in COLUNAS_PERDAS if c in df.columns]
        if not col_perdas_reais: return "null"
        df_valid = df[df[col_perdas_reais].notna().any(axis=1)].copy()
        df_valid['TOTAL_PERDAS'] = df_valid[col_perdas_reais].fillna(0).sum(axis=1)
        if c_turno in df_valid.columns: df_valid = df_valid[df_valid[c_turno].isin(TURNOS_VALIDOS)]
        if len(df_valid)==0: return "null"
        date_col = next((c for c in df_valid.columns if 'data' in str(c).lower()), None)
        if date_col:
            df_valid[date_col] = pd.to_datetime(df_valid[date_col], errors='coerce')
            def format_week(d):
                if pd.isnull(d): return "S/ Data"
                try:
                    wk=d.isocalendar()[1]; start=d-pd.Timedelta(days=d.weekday()); end=start+pd.Timedelta(days=6)
                    return f"Sem. {wk:02d} ({start.strftime('%d/%m')} a {end.strftime('%d/%m')})"
                except: return "S/ Data"
            df_valid['SEMANA'] = df_valid[date_col].apply(format_week)
        else: df_valid['SEMANA']="S/ Data"
        faz_col = next((c for c in df_valid.columns if 'fazenda' in str(c).lower() or 'propriedade' in str(c).lower()), None)
        df_valid['FAZENDA'] = df_valid[faz_col].fillna('S/ Fazenda').astype(str).str.strip() if faz_col else "S/ Fazenda"
        zona_col = next((c for c in df_valid.columns if 'zona' in str(c).lower()), None)
        df_valid['ZONA'] = df_valid[zona_col].fillna('S/ Zona').astype(str).str.strip() if zona_col else "S/ Zona"
        def media_fill(x): return x.fillna(0).mean()
        agrupadores = [c for c in [c_equip,c_frente,c_turno,'SEMANA','FAZENDA','ZONA'] if c in df_valid.columns]
        detail = df_valid.groupby(agrupadores).agg(
            media=('TOTAL_PERDAS','mean'), count=('TOTAL_PERDAS','count'),
            agr_touceira=('AGR_TOUCEIRA',media_fill), agr_palmito=('AGR_PALMITO',media_fill),
            cana_inteira=('CANA_INTEIRA',media_fill), rebolo=('REBOLO',media_fill),
            pedaco=('PEDACO',media_fill), estilhaco=('ESTILHACO',media_fill), lasca=('LASCA',media_fill),
        ).reset_index()
        renames = {c_equip:'equip',c_frente:'frente',c_turno:'turno','SEMANA':'semana','FAZENDA':'fazenda','ZONA':'zona'}
        detail = detail.rename(columns=renames)
        frentes_limpas = sorted(set(int(extrair_frente_valida(x)) for x in df_valid[c_frente].unique() if extrair_frente_valida(x)))
        semanas_unicas = sorted([str(x) for x in df_valid['SEMANA'].unique() if str(x) not in ('NaT','nan')])
        fazendas_unicas = sorted([str(x) for x in df_valid['FAZENDA'].unique() if str(x) not in ('nan','S/ Fazenda')])
        zonas_unicas = sorted([str(x) for x in df_valid['ZONA'].unique() if str(x) not in ('nan','S/ Zona')])
        data = {'raw':detail.round(3).to_dict('records'),'frentes':frentes_limpas,'semanas':semanas_unicas,'fazendas':fazendas_unicas,'zonas':zonas_unicas}
        def conv(obj):
            if isinstance(obj,(np.integer,)): return int(obj)
            if isinstance(obj,(np.floating,)): return float(obj)
            if isinstance(obj,np.ndarray): return obj.tolist()
            return obj
        return json.dumps(data, default=conv, separators=(',',':'))
    except Exception as e:
        print(f"  [AVISO] Erro perdas: {e}"); return "null"

# ============================================================
# EXTRATOR 3: OVERVIEW CCT (REESCRITO - Lê todas as abas de apoio)
# ============================================================
def extract_overview(caminho_metricas):
    if not os.path.exists(caminho_metricas): return "null"
    try:
        wb = load_workbook(caminho_metricas, data_only=True)
        sheets = wb.sheetnames
        print(f"  [OK] Métricas: {len(sheets)} abas encontradas")

        # ── 1) KPIs de Apoio_KPI_Frente ──
        kpis = {'todas':{'tcd':{'meta':0,'real':0},'corte':[],'carreg':[],'transp':[]},}
        if 'Apoio_KPI_Frente' in sheets:
            ws = wb['Apoio_KPI_Frente']
            area_map = {'TCD':'tcd','Corte':'corte','Carregamento':'carreg','Transporte':'transp'}
            lower_better = {'Manobra','T2','Raio','Imp. Vegetal'}
            for r in range(2, ws.max_row+1):
                frente_raw = ws.cell(r,1).value
                name = ws.cell(r,2).value
                meta = sf(ws.cell(r,3).value)
                real = sf(ws.cell(r,4).value)
                area = ws.cell(r,6).value
                lb = ws.cell(r,7).value == 'SIM'
                if not name: continue
                fr_key = 'todas' if str(frente_raw).strip().lower()=='total' else str(frente_raw).replace('.0','')
                if fr_key not in kpis:
                    kpis[fr_key] = {'tcd':{'meta':0,'real':0},'corte':[],'carreg':[],'transp':[]}
                cat = area_map.get(area)
                if not cat: continue
                if cat == 'tcd':
                    kpis[fr_key]['tcd'] = {'meta':meta,'real':real}
                else:
                    short = name.split('[')[0].strip()
                    unit_map = {'ton/viag':'t/viag','ton/maq':'t/m/d','km/h':'km/h','km':'km','min':'min','%':'%'}
                    u = ''
                    for k,v in unit_map.items():
                        if k in name.lower(): u=v; break
                    tab_map = {'corte':4,'transp':5}
                    kpis[fr_key][cat].append({'name':short,'meta':meta,'real':real,'unit':u,'higher_is_better':not lb,'tab':tab_map.get(cat)})

        frentes_list = ['todas'] + sorted([k for k in kpis if k != 'todas'], key=lambda x: int(x) if x.isdigit() else 999)

        # ── 2) Fleet detail from Apoio_Frota_Detalhe (agregado) ──
        fleet = []
        if 'Apoio_Frota_Detalhe' in sheets:
            ws = wb['Apoio_Frota_Detalhe']
            for r in range(2, ws.max_row+1):
                eq = ws.cell(r,1).value
                fr = ws.cell(r,2).value
                if not eq: continue
                totalH = sf(ws.cell(r,14).value)
                if totalH <= 0: continue
                fleet.append({
                    'equip':str(eq),
                    'frente':str(fr).replace('.0','') if fr else '',
                    'velVazio':sf(ws.cell(r,4).value),
                    'velCarreg':sf(ws.cell(r,5).value),
                    't1':sf(ws.cell(r,6).value),
                    't2':sf(ws.cell(r,7).value),
                    'raio':sf(ws.cell(r,8).value),
                    'densidade':sf(ws.cell(r,9).value),
                    'manut':sf(ws.cell(r,10).value),
                    'agTransb':sf(ws.cell(r,11).value),
                    'semApont':sf(ws.cell(r,12).value),
                    'patio':sf(ws.cell(r,13).value),
                    'totalH':totalH,
                    'pctOfensor':round(sf(ws.cell(r,15).value)*100,1),
                })

        # ── 2b) Raw trip rows with dates (for JS date filtering) ──
        raw_trips = []
        if 'Apoio_Tempo_Cargas' in sheets:
            ws = wb['Apoio_Tempo_Cargas']
            for r in range(2, ws.max_row+1):
                eq = ws.cell(r,1).value
                dt = ws.cell(r,2).value
                if not eq or not dt: continue
                raw_trips.append({
                    'eq':str(eq),'dt':dt.strftime('%Y-%m-%d') if hasattr(dt,'strftime') else str(dt)[:10],
                    'fr':str(ws.cell(r,3).value).replace('.0','') if ws.cell(r,3).value else '',
                    't1':sf(ws.cell(r,10).value,1),'t2':sf(ws.cell(r,11).value,1),
                    'dens':sf(ws.cell(r,21).value,2),'raio':sf(ws.cell(r,23).value,2),
                })

        # ── 2c) Raw bordo rows with dates ──
        raw_bordo = []
        if 'Apoio_Bordo' in sheets:
            ws = wb['Apoio_Bordo']
            for r in range(2, ws.max_row+1):
                eq = ws.cell(r,1).value
                dt = ws.cell(r,2).value
                if not eq or not dt: continue
                raw_bordo.append({
                    'eq':str(eq),'dt':dt.strftime('%Y-%m-%d') if hasattr(dt,'strftime') else str(dt)[:10],
                    'vz':sf(ws.cell(r,4).value,2),'cr':sf(ws.cell(r,5).value,2),
                    'man':sf(ws.cell(r,11).value,2)+sf(ws.cell(r,12).value,2)+sf(ws.cell(r,13).value,2),
                    'ag':sf(ws.cell(r,8).value,2),'sa':sf(ws.cell(r,14).value,2),
                    'pat':sf(ws.cell(r,9).value,2),'tot':sf(ws.cell(r,16).value,2),
                    'vpv':sf(ws.cell(r,17).value,0),'hov':sf(ws.cell(r,18).value,0),
                    'vpc':sf(ws.cell(r,20).value,0),'hoc':sf(ws.cell(r,21).value,0),
                })

        # ── 3) Ciclo Logístico diário ──
        ciclo = []
        if 'Apoio_Ciclo_Logistico' in sheets:
            ws = wb['Apoio_Ciclo_Logistico']
            for r in range(2, ws.max_row+1):
                dt = ws.cell(r,1).value
                if dt is None: continue
                if isinstance(dt, str) and dt.strip().upper()=='TOTAL': continue
                ciclo.append({
                    'data': dt.strftime('%d/%m') if hasattr(dt,'strftime') else str(dt),
                    'viagens':sf(ws.cell(r,2).value,0),
                    'raioMeta':sf(ws.cell(r,3).value,1),'raioReal':sf(ws.cell(r,4).value,1),
                    'densMeta':sf(ws.cell(r,5).value,1),'densReal':sf(ws.cell(r,6).value,1),
                    'velVMeta':sf(ws.cell(r,7).value,1),'velVReal':sf(ws.cell(r,8).value,1),
                    'velCMeta':sf(ws.cell(r,9).value,1),'velCReal':sf(ws.cell(r,10).value,1),
                    't2Meta':sf(ws.cell(r,11).value,0),'t2Real':sf(ws.cell(r,12).value,0),
                    't4Real':sf(ws.cell(r,13).value,0),
                    'patio':sf(ws.cell(r,14).value,1),'manut':sf(ws.cell(r,15).value,1),
                    'agTransb':sf(ws.cell(r,16).value,1),'semApont':sf(ws.cell(r,17).value,1),
                    'totalH':sf(ws.cell(r,18).value,1),
                    'pctOf':round(sf(ws.cell(r,19).value)*100,1),
                })

        # ── 4) Velocidade por Frente ──
        vel_fr = []
        if 'Apoio_Vel_Frente' in sheets:
            ws = wb['Apoio_Vel_Frente']
            for r in range(2, ws.max_row+1):
                fr = ws.cell(r,1).value
                if not fr: continue
                vel_fr.append({
                    'frente':str(fr).replace('.0',''),
                    'velVReal':sf(ws.cell(r,2).value,1),'velCReal':sf(ws.cell(r,3).value,1),
                    'velVMeta':sf(ws.cell(r,4).value,1),'velCMeta':sf(ws.cell(r,5).value,1),
                    'densReal':sf(ws.cell(r,6).value,1),'densMeta':sf(ws.cell(r,7).value,1),
                    't2Real':sf(ws.cell(r,8).value,0),'t2Meta':sf(ws.cell(r,9).value,0),
                    'raioReal':sf(ws.cell(r,10).value,1),'raioMeta':sf(ws.cell(r,11).value,1),
                    'viagens':sf(ws.cell(r,12).value,0),'volume':sf(ws.cell(r,13).value,0),
                })

        # ── 4b) Corte por Frente from Apoio_Corte_Frente ──
        corte_fr = []
        if 'Apoio_Corte_Frente' in sheets:
            ws = wb['Apoio_Corte_Frente']
            for r in range(2, ws.max_row+1):
                fr = ws.cell(r,1).value
                if not fr: continue
                corte_fr.append({
                    'frente':str(fr).replace('.0',''),
                    'corteH':sf(ws.cell(r,3).value,1),
                    'manobraH':sf(ws.cell(r,4).value,1),
                    'deslocH':sf(ws.cell(r,5).value,1),
                    'chuvaH':sf(ws.cell(r,6).value,1),
                    'limitIndH':sf(ws.cell(r,7).value,1),
                    'manutH':sf(ws.cell(r,8).value,1),
                    'patioH':sf(ws.cell(r,9).value,1),
                    'semApontH':sf(ws.cell(r,10).value,1),
                    'totalH':sf(ws.cell(r,11).value,1),
                    'velCorte':sf(ws.cell(r,12).value,2),
                    'pctCorte':round(sf(ws.cell(r,13).value)*100,1),
                    'pctManobra':round(sf(ws.cell(r,14).value)*100,1),
                    'pctOfensores':round(sf(ws.cell(r,15).value)*100,1),
                    'hrsEfMeta':sf(ws.cell(r,16).value,1),
                    'hrsEfReal':sf(ws.cell(r,17).value,1),
                    'atr':sf(ws.cell(r,18).value,1),
                    'chuvaHMaq':sf(ws.cell(r,19).value,1),
                    'manobraMeta':round(sf(ws.cell(r,20).value)*100,1),
                    'manobraReal':round(sf(ws.cell(r,21).value)*100,1),
                })
            print(f"  [OK] Corte/Frente: {len(corte_fr)} frentes")

        # ── 4c) Raw bordo CD rows for date filtering ──
        raw_bordo_cd = []
        if 'Apoio_Bordo_CD' in sheets:
            ws = wb['Apoio_Bordo_CD']
            for r in range(2, ws.max_row+1):
                frota = ws.cell(r,1).value
                dt = ws.cell(r,2).value
                if not frota or not dt: continue
                raw_bordo_cd.append({
                    'fr':str(ws.cell(r,4).value).replace('.0','') if ws.cell(r,4).value else '',
                    'dt':dt.strftime('%Y-%m-%d') if hasattr(dt,'strftime') else str(dt)[:10],
                    'eq':str(int(float(frota))) if frota else '',
                    'co':sf(ws.cell(r,5).value,2),'mn':sf(ws.cell(r,6).value,2),
                    'ds':sf(ws.cell(r,7).value,2),'ch':sf(ws.cell(r,9).value,2),
                    'li':sf(ws.cell(r,10).value,2),'ma':sf(ws.cell(r,11).value,2),
                    'pa':sf(ws.cell(r,12).value,2),'sa':sf(ws.cell(r,13).value,2),
                    'tot':sf(ws.cell(r,16).value,2),
                    'vpC':sf(ws.cell(r,17).value,0),'hoC':sf(ws.cell(r,18).value,0),
                })
            print(f"  [OK] Bordo CD: {len(raw_bordo_cd)} rows")

        # ── 4d) Operator-level data from Base bordo CD raw ──
        raw_operadores_cd = []
        if 'Base bordo CD' in sheets:
            ws = wb['Base bordo CD']
            # Encontrar dinamicamente a coluna de Data
            date_col = 2
            for c in range(1, min(ws.max_column+1, 20)):
                h_val = str(ws.cell(1, c).value).lower() if ws.cell(1, c).value else ''
                if 'data' in h_val or 'dt' in h_val:
                    date_col = c
                    break

            for r in range(2, ws.max_row+1):
                dt_val = ws.cell(r, date_col).value
                op = ws.cell(r,3).value  # CD_OPERADOR
                frota = ws.cell(r,4).value  # Frota
                frente = ws.cell(r,5).value  # Frente
                desc_op = ws.cell(r,12).value  # DESC_OPERAC
                hrs = ws.cell(r,13).value  # CB (h)
                vel_p = ws.cell(r,15).value  # VEL_POND
                hr_op = ws.cell(r,17).value  # HR_OPERACIONAIS_VEL
                
                if not op or not frota: continue
                fr_num = str(frente).split('-')[-1] if frente and '-' in str(frente) else ''
                dt_str = dt_val.strftime('%Y-%m-%d') if hasattr(dt_val, 'strftime') else str(dt_val)[:10] if dt_val else ''

                raw_operadores_cd.append({
                    'dt': dt_str,
                    'op': str(int(float(op))) if op else '',
                    'eq': str(int(float(frota))) if frota else '',
                    'fr': fr_num,
                    'desc': str(desc_op) if desc_op else '',
                    'h': sf(hrs,3),
                    'vpC': sf(vel_p,0),'hoC': sf(hr_op,0),
                })
            print(f"  [OK] Operadores CD: {len(raw_operadores_cd)} rows")

        # ── 2d) Collect all available dates (Movemos para o fim para pegar todos) ──
        all_dates = sorted(set(r.get('dt') for r in raw_trips if r.get('dt')) | 
                           set(r.get('dt') for r in raw_bordo if r.get('dt')) | 
                           set(r.get('dt') for r in raw_bordo_cd if r.get('dt')) | 
                           set(r.get('dt') for r in raw_operadores_cd if r.get('dt')))

        # ── 5) Metas de referência (Apoio_Metas_Transporte) ──
        metas = {'velVazio':36.6,'velCarreg':27.0,'raio':29.4,'t1':90,'t2':27,'t4':69.2,'densidade':33.5,'pctOfensor':96.0}
        if 'Apoio_Metas_Transporte' in sheets:
            wm = wb['Apoio_Metas_Transporte']
            # A2=Raio, B2=Densidade, C2=Vel.Vazio, D2=Vel.Carreg, E2=T2, F2=T4, G2=%Ofensores
            metas['raio'] = sf(wm.cell(2,1).value,1) or 29.4
            metas['densidade'] = sf(wm.cell(2,2).value,1) or 33.5
            metas['velVazio'] = sf(wm.cell(2,3).value,1) or 36.6
            metas['velCarreg'] = sf(wm.cell(2,4).value,1) or 27.0
            metas['t2'] = sf(wm.cell(2,5).value,0) or 27
            metas['t4'] = sf(wm.cell(2,6).value,0) or 69
            metas['pctOfensor'] = round(sf(wm.cell(2,7).value)*100, 0) or 96
            print(f"  [OK] Metas Transporte: VV={metas['velVazio']} VC={metas['velCarreg']} Raio={metas['raio']} Dens={metas['densidade']} T2={metas['t2']} T4={metas['t4']} %Of={metas['pctOfensor']}")
        elif 'KPI por frente' in sheets:
            wk = wb['KPI por frente']
            metas['velVazio'] = sf(wk.cell(19,8).value,1) or 36.6
            metas['velCarreg'] = sf(wk.cell(20,8).value,1) or 27.0
            metas['raio'] = sf(wk.cell(25,8).value,1) or 29.4
            metas['t2'] = sf(wk.cell(24,8).value,0) or 27
            metas['densidade'] = sf(wk.cell(7,8).value,1) or 33.5

        # ── 6) Bordo ranking ──
        bordo = []
        if 'Apoio_Bordo' in sheets:
            ws = wb['Apoio_Bordo']
            bagg = defaultdict(lambda:{'vazio':0,'carreg':0,'total':0})
            for r in range(2, ws.max_row+1):
                eq = ws.cell(r,1).value
                if not eq: continue
                k = str(eq)
                bagg[k]['vazio'] += sf(ws.cell(r,4).value)
                bagg[k]['carreg'] += sf(ws.cell(r,5).value)
                bagg[k]['total'] += sf(ws.cell(r,16).value)
            eq_fr = {f['equip']:f['frente'] for f in fleet}
            for eq, v in bagg.items():
                if v['total'] > 0:
                    pv = round(v['vazio']/v['total']*100,1)
                    pp = round((1-(v['vazio']+v['carreg'])/v['total'])*100,1)
                    bordo.append({'equip':eq,'frente':eq_fr.get(eq,''),'pctVazio':pv,'pctParado':pp,'pctOfensor':pp})

        # Build per-frente bordo
        for fr in frentes_list:
            if fr not in kpis: continue
            if fr == 'todas':
                kpis[fr]['bordo'] = sorted(bordo, key=lambda x:-x['pctOfensor'])[:12]
            else:
                kpis[fr]['bordo'] = sorted([b for b in bordo if b['frente']==fr], key=lambda x:-x['pctOfensor'])

        wb.close()

        result = {
            'kpis': kpis,
            'frentes': frentes_list,
            'fleet': fleet,
            'ciclo': ciclo,
            'vel_fr': vel_fr,
            'metas': metas,
            'raw_trips': raw_trips,
            'raw_bordo': raw_bordo,
            'dates': all_dates,
            'corte_fr': corte_fr,
            'raw_bordo_cd': raw_bordo_cd,
            'raw_op_cd': raw_operadores_cd,
        }
        print(f"  [OK] KPIs: {sum(len(v.get('corte',[])) + len(v.get('carreg',[])) + len(v.get('transp',[])) for v in kpis.values())} indicadores")
        print(f"  [OK] Fleet: {len(fleet)} equips | Ciclo: {len(ciclo)} dias | Trips: {len(raw_trips)} | Bordo: {len(raw_bordo)} rows")
        return json.dumps(result, separators=(',',':'), ensure_ascii=False)

    except Exception as e:
        print(f"  [ERRO] extract_overview: {e}")
        traceback.print_exc()
        return "null"


# ============================================================
# MONTADOR HTML
# ============================================================
def build_html(json_prod, frente_names, json_perdas, json_overview):
    n_fr = len(frente_names)
    fr_list_js = json.dumps(frente_names)
    fc_obj = '{' + ','.join(f"'{fn}':{json.dumps(FRENTE_COLORS[i%len(FRENTE_COLORS)][0])}" for i,fn in enumerate(frente_names)) + '}'
    fcd_obj = '{' + ','.join(f"'{fn}':{json.dumps(FRENTE_COLORS[i%len(FRENTE_COLORS)][1])}" for i,fn in enumerate(frente_names)) + '}'
    grid_cols = f"repeat({min(n_fr,4)},1fr)"
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M")

    html = f'''<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>CCT Ipaussu — Painel Integrado | Raízen</title>
<link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500;600;700&display=swap" rel="stylesheet">
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.2.0"></script>
<style>
:root {{
    --rz:#00843D; --rzl:#00A84E; --rzd:#006830; --rzg:rgba(0,132,61,.12); --rzg2:rgba(0,132,61,.06);
    --bg:#f5f7f6; --s1:#fff; --s2:#f0f4f2; --bd:#e0e8e3; --tx:#1a2e22; --tx2:#5a7265;
    --red:#d63031; --redd:rgba(214,48,49,.1); --am:#e17055; --amd:rgba(225,112,85,.1);
    --bl:#0984e3; --cy:#00b894; --pr:#6c5ce7;
    --sh:0 1px 3px rgba(0,40,20,.06),0 1px 2px rgba(0,40,20,.04); --sh2:0 4px 12px rgba(0,40,20,.08);
    --p-bg:var(--bg); --p-surface:var(--s1); --p-surface-2:var(--s2); --p-surface-3:var(--bd);
    --p-border:var(--bd); --p-text:var(--tx); --p-text-muted:var(--tx2); --p-text-dim:#78909c;
    --p-green:var(--rz); --p-green-bg:var(--rzg); --p-green-border:rgba(0,132,61,0.25);
    --p-red:var(--red); --p-red-bg:var(--redd); --p-red-border:rgba(214,48,49,0.25);
    --p-yellow:var(--am); --p-yellow-bg:var(--amd); --p-yellow-border:rgba(225,112,85,0.25);
    --p-accent:var(--bl); --p-accent-bg:rgba(9,132,227,0.1);
    --ok:#27ae60; --danger:#d63031; --warn:#f39c12;
}}
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Plus Jakarta Sans',sans-serif;background:var(--bg);color:var(--tx);min-height:100vh;-webkit-font-smoothing:antialiased}}
.app{{max-width:1500px;margin:0 auto;padding:16px 20px}}
.hdr{{display:flex;justify-content:space-between;align-items:center;padding:16px 24px;background:var(--rz);border-radius:14px;margin-bottom:16px;box-shadow:var(--sh2)}}
.hdr-left{{display:flex;align-items:center;gap:14px}}.hdr-logo{{width:36px;height:36px;background:#fff;border-radius:8px;display:flex;align-items:center;justify-content:center;font-weight:800;color:var(--rz);font-size:14px}}
.hdr h1{{font-size:18px;font-weight:700;color:#fff}}.hdr h1 span{{opacity:.7;font-weight:500}}
.hdr-r{{display:flex;gap:10px}}.badge{{padding:5px 12px;border-radius:20px;font-size:11px;font-weight:600;font-family:'JetBrains Mono',monospace;background:rgba(255,255,255,.2);color:#fff}}
.tabs{{display:flex;gap:4px;margin-bottom:16px;background:var(--s1);padding:4px;border-radius:12px;box-shadow:var(--sh);width:fit-content;overflow-x:auto}}
.tab{{padding:10px 24px;border-radius:10px;border:none;background:transparent;color:var(--tx2);cursor:pointer;font-size:13px;font-weight:600;font-family:'Plus Jakarta Sans',sans-serif;transition:all .2s;white-space:nowrap}}
.tab:hover{{background:var(--rzg);color:var(--rz)}}.tab.on{{background:var(--rz);color:#fff;box-shadow:0 2px 8px rgba(0,132,61,.3)}}
.filters{{display:flex;gap:6px;margin-bottom:16px;flex-wrap:wrap}}
.fb{{padding:6px 14px;border:1.5px solid var(--bd);border-radius:8px;background:var(--s1);color:var(--tx2);cursor:pointer;font-size:12px;font-weight:600;font-family:'Plus Jakarta Sans',sans-serif}}
.fb:hover{{border-color:var(--rz);color:var(--rz)}}.fb.on{{background:var(--rz);border-color:var(--rz);color:#fff}}
.kg{{display:grid;grid-template-columns:repeat(auto-fit,minmax(185px,1fr));gap:12px;margin-bottom:16px}}
.k{{background:var(--s1);border:1px solid var(--bd);border-radius:12px;padding:16px 18px;position:relative;overflow:hidden;box-shadow:var(--sh)}}
.k::before{{content:'';position:absolute;top:0;left:0;width:4px;height:100%;border-radius:4px 0 0 4px}}
.k.cg::before{{background:var(--rz)}}.k.cb::before{{background:var(--bl)}}.k.ca::before{{background:var(--am)}}.k.cc::before{{background:var(--cy)}}.k.cp::before{{background:var(--pr)}}.k.cr::before{{background:var(--red)}}
.kl{{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.8px;color:var(--tx2);margin-bottom:6px}}
.kv{{font-family:'JetBrains Mono',monospace;font-size:24px;font-weight:700}}.ks{{font-size:10px;color:var(--tx2);margin-top:3px;font-family:'JetBrains Mono',monospace}}
.kp{{display:inline-block;padding:3px 8px;border-radius:6px;font-size:10px;font-weight:700;font-family:'JetBrains Mono',monospace;margin-top:6px}}
.pu{{background:var(--rzg);color:var(--rz)}}.pd{{background:var(--redd);color:var(--red)}}
.cg2{{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:16px}}
.cc{{background:var(--s1);border:1px solid var(--bd);border-radius:12px;padding:18px;box-shadow:var(--sh)}}.cc.full{{grid-column:1/-1}}
.ct{{font-size:13px;font-weight:700;margin-bottom:14px;display:flex;align-items:center;gap:8px}}.ct .dot{{width:8px;height:8px;border-radius:50%}}
.cw{{position:relative;height:260px}}.cw.tall{{height:320px}}
.tc{{background:var(--s1);border:1px solid var(--bd);border-radius:12px;padding:18px;margin-bottom:16px;overflow-x:auto;box-shadow:var(--sh)}}
.tc table{{width:100%;border-collapse:collapse;font-size:11px}}
.tc th{{text-align:center;padding:10px;color:var(--tx2);font-weight:700;text-transform:uppercase;font-size:9px;border-bottom:2px solid var(--bd);white-space:nowrap}}
.tc td{{padding:8px 10px;border-bottom:1px solid var(--s2);font-family:'JetBrains Mono',monospace;font-size:11px;white-space:nowrap;text-align:center}}
.tc tr:hover td{{background:var(--rzg2)}}.vp{{color:var(--rz);font-weight:600}}.vn{{color:var(--red);font-weight:600}}.vw{{color:var(--am);font-weight:600}}
.frente-grid{{display:grid;grid-template-columns:{grid_cols};gap:12px;margin-bottom:16px}}
.fc{{background:var(--s1);border:1px solid var(--bd);border-radius:12px;padding:16px;position:relative;overflow:hidden;box-shadow:var(--sh)}}
.fc-name{{font-size:13px;font-weight:800;margin-bottom:10px;display:flex;align-items:center;gap:6px}}.fc-name .dot{{width:10px;height:10px;border-radius:50%}}
.fc-row{{display:flex;justify-content:space-between;padding:5px 0;border-bottom:1px solid var(--s2);font-size:11px}}.fc-row:last-child{{border-bottom:none}}
.fc-row .label{{color:var(--tx2)}}.fc-row .val{{font-family:'JetBrains Mono',monospace;font-weight:600}}
.gap-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:16px}}
.gap-card{{background:var(--s1);border:1px solid var(--bd);border-radius:12px;padding:18px;box-shadow:var(--sh)}}.gap-card h3{{font-size:13px;font-weight:800;margin-bottom:14px}}
.gap-bar-row{{display:flex;align-items:center;gap:10px;margin-bottom:10px}}.gap-label{{width:40px;font-size:11px;font-weight:700;color:var(--tx2);text-align:right}}
.gap-bar-wrap{{flex:1;height:30px;background:var(--s2);border-radius:8px;overflow:hidden}}
.gap-bar{{height:100%;border-radius:8px;display:flex;align-items:center;justify-content:flex-end;padding-right:10px;font-size:10px;font-weight:700;font-family:'JetBrains Mono',monospace;color:#fff;min-width:45px}}
.gap-val{{width:70px;font-size:11px;font-family:'JetBrains Mono',monospace;font-weight:600;text-align:right}}
.gap-diff{{font-size:10px;font-family:'JetBrains Mono',monospace}}.gap-diff.pos{{color:var(--rz)}}.gap-diff.neg{{color:var(--red)}}
.section-title{{font-size:16px;font-weight:800;margin:22px 0 14px;display:flex;align-items:center;gap:8px}}.section-title .bar{{width:4px;height:22px;border-radius:2px;background:var(--rz)}}
/* Perdas styles */
.perdas-wrapper{{background:transparent;border-radius:14px}}
.p-filters{{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:24px;align-items:center}}
.p-filter-btn{{padding:7px 18px;border-radius:6px;border:1px solid var(--p-border);background:var(--p-surface);color:var(--p-text-muted);cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;font-size:13px;font-weight:600;transition:all .2s;outline:none}}
.p-filter-btn:hover{{border-color:var(--p-text-dim);color:var(--p-text)}}.p-filter-btn.active{{background:var(--p-accent-bg);border-color:var(--p-accent);color:var(--p-accent)}}
.p-filter-label{{display:flex;align-items:center;font-size:12px;color:var(--p-text-dim);text-transform:uppercase;letter-spacing:1px;font-weight:700;margin-right:4px}}
select.p-filter-btn{{appearance:none;padding-right:30px;background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='8'%3E%3Cpath d='M1 1l5 5 5-5' stroke='%235a7265' fill='none' stroke-width='2' stroke-linecap='round'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:right 10px center}}
.p-kpi-row{{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:14px;margin-bottom:24px}}
.p-kpi-card{{background:var(--p-surface);border:1px solid var(--p-border);border-radius:10px;padding:18px 20px;position:relative;overflow:hidden;box-shadow:var(--sh)}}
.p-kpi-card::before{{content:'';position:absolute;top:0;left:0;right:0;height:3px}}
.p-kpi-card.status-green::before{{background:var(--p-green)}}.p-kpi-card.status-red::before{{background:var(--p-red)}}.p-kpi-card.status-yellow::before{{background:var(--p-yellow)}}
.p-kpi-label{{font-size:11px;color:var(--p-text-dim);text-transform:uppercase;letter-spacing:1px;font-weight:700}}
.p-kpi-value{{font-family:'JetBrains Mono',monospace;font-size:28px;font-weight:700;margin:6px 0 2px}}.p-kpi-unit{{font-size:14px;color:var(--p-text-muted);font-weight:400}}
.p-kpi-status{{font-size:12px;margin-top:4px}}
.p-grid-2{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:24px}}
.p-card{{background:var(--p-surface);border:1px solid var(--p-border);border-radius:10px;overflow:hidden;margin-bottom:24px;box-shadow:var(--sh)}}
.p-card-header{{padding:14px 18px;border-bottom:1px solid var(--p-border);display:flex;align-items:center;justify-content:space-between}}
.p-card-title{{font-size:14px;font-weight:700;color:var(--tx)}}.p-card-body{{padding:16px 18px}}
.p-table-wrap{{overflow-x:auto}}.p-table{{width:100%;border-collapse:collapse}}
.p-table th{{text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:1px;color:var(--p-text-dim);padding:10px;border-bottom:2px solid var(--p-border);font-weight:700;white-space:nowrap}}
.p-table td{{padding:10px;border-bottom:1px solid var(--p-surface-2);font-size:13px;color:var(--p-text);white-space:nowrap}}
.p-table tr:last-child td{{border-bottom:none}}.p-table tr:hover td{{background:rgba(0,132,61,0.03)}}
.p-mono{{font-family:'JetBrains Mono',monospace;font-size:12px}}
.p-text-green{{color:var(--p-green)}}.p-text-red{{color:var(--p-red)}}.p-text-yellow{{color:var(--p-yellow)}}.p-text-muted{{color:var(--p-text-muted)}}
.p-status-pill{{display:inline-flex;align-items:center;gap:4px;padding:3px 10px;border-radius:99px;font-size:11px;font-weight:700}}
.p-pill-green{{background:var(--p-green-bg);color:var(--p-green);border:1px solid var(--p-green-border)}}
.p-pill-red{{background:var(--p-red-bg);color:var(--p-red);border:1px solid var(--p-red-border)}}
.p-pill-yellow{{background:var(--p-yellow-bg);color:var(--p-yellow);border:1px solid var(--p-yellow-border)}}
.p-bar-cell{{display:flex;align-items:center;gap:8px}}.p-bar-track{{flex:1;height:6px;background:var(--p-surface-3);border-radius:3px;overflow:hidden}}.p-bar-fill{{height:100%;border-radius:3px;transition:width .4s ease}}
.p-chart-container{{position:relative;height:260px}}
.p-breakdown-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:10px}}
.p-breakdown-item{{background:var(--p-surface-2);border-radius:8px;padding:12px 14px;border:1px solid var(--p-border)}}
.p-breakdown-label{{font-size:10px;color:var(--p-text-dim);text-transform:uppercase;letter-spacing:0.5px;margin-bottom:4px;font-weight:700}}
.p-breakdown-value{{font-family:'JetBrains Mono',monospace;font-size:18px;font-weight:700;color:var(--tx)}}
.p-breakdown-bar{{height:3px;border-radius:2px;margin-top:8px;background:var(--p-surface-3);overflow:hidden}}.p-breakdown-bar-fill{{height:100%;border-radius:2px;background:var(--p-accent)}}
.p-rank-num{{width:24px;height:24px;border-radius:6px;display:inline-flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;font-family:'JetBrains Mono',monospace}}
.p-rank-1{{background:var(--p-red-bg);color:var(--p-red);border:1px solid var(--p-red-border)}}
.p-rank-2{{background:var(--p-red-bg);color:var(--p-red);border:1px solid var(--p-red-border)}}
.p-rank-3{{background:var(--p-yellow-bg);color:var(--p-yellow);border:1px solid var(--p-yellow-border)}}
.p-rank-default{{background:var(--p-surface-3);color:var(--p-text-muted)}}
/* Overview CCT */
.ov-panel{{background:var(--s1);border:1px solid var(--bd);border-radius:12px;padding:18px;box-shadow:var(--sh);margin-bottom:16px}}
.ov-panel-title{{font-size:13px;font-weight:700;margin-bottom:14px;display:flex;align-items:center;gap:8px;color:var(--tx2);text-transform:uppercase;letter-spacing:.06em}}
.ov-panel-title .dot{{width:8px;height:8px;border-radius:50%;flex-shrink:0}}
.ov-tcd{{background:var(--s1);border-radius:16px;padding:24px;box-shadow:var(--sh2);margin-bottom:20px;display:grid;grid-template-columns:1fr auto;align-items:center;gap:24px;position:relative;overflow:hidden;border:1px solid var(--bd)}}
.ov-tcd::before{{content:'';position:absolute;top:0;left:0;width:6px;height:100%}}
.ov-tcd.ok::before{{background:var(--ok)}}.ov-tcd.bad::before{{background:var(--danger)}}
.ov-tcd-lbl{{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.1em;color:var(--tx2);margin-bottom:4px}}
.ov-tcd-ttl{{font-size:15px;font-weight:700;margin-bottom:12px}}
.ov-tcd-vals{{display:flex;gap:24px;align-items:baseline}}
.ov-tcd-val{{display:flex;flex-direction:column}}.ov-tcd-val span:first-child{{font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.08em;color:var(--tx2)}}
.ov-tcd-val strong{{font-size:36px;font-weight:800;font-family:'JetBrains Mono',monospace;letter-spacing:-.03em}}
.ov-tcd-val.meta strong{{color:var(--tx2);font-weight:600;font-size:20px}}
.ov-tcd-badge{{padding:12px 24px;border-radius:14px;font-size:16px;font-weight:800;text-align:center;white-space:nowrap}}
.ov-tcd-badge.ok{{background:var(--rzg);color:var(--rz);border:2px solid rgba(0,132,61,.2)}}
.ov-tcd-badge.bad{{background:var(--redd);color:var(--red);border:2px solid rgba(214,48,49,.2)}}
.ov-cct-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:16px}}
.ov-col-hdr{{padding:10px 14px;border-radius:8px;font-weight:800;font-size:13px;margin-bottom:8px;display:flex;align-items:center;gap:8px}}
.ov-col-hdr.corte{{background:#e8f8ef;color:#00843D}}.ov-col-hdr.carreg{{background:#e3f0ff;color:#0984e3}}.ov-col-hdr.transp{{background:#fff4e0;color:#d48806}}
.ov-kpi{{background:var(--s1);border:1px solid var(--bd);border-radius:8px;padding:12px 14px;box-shadow:var(--sh);border-left:4px solid var(--bd);margin-bottom:6px;display:grid;grid-template-columns:1fr auto;align-items:center;gap:6px}}
.ov-kpi.ok{{border-left-color:var(--ok)}}.ov-kpi.bad{{border-left-color:var(--danger)}}
.ov-kpi-nm{{font-size:11px;font-weight:600;color:var(--tx2)}}.ov-kpi-vl{{font-size:18px;font-weight:800;font-family:'JetBrains Mono',monospace}}
.ov-kpi.bad .ov-kpi-vl{{color:var(--danger)}}.ov-kpi-mt{{font-size:10px;color:var(--tx2)}}.ov-kpi-mt b{{font-weight:700}}
.ov-kpi-ic{{width:28px;height:28px;border-radius:50%;display:grid;place-items:center;font-size:12px}}
.ov-kpi.ok .ov-kpi-ic{{background:var(--rzg)}}.ov-kpi.bad .ov-kpi-ic{{background:var(--redd)}}
/* Gauge row */
.ov-gauge-row{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:16px}}
.ov-gauge{{background:var(--s1);border:1px solid var(--bd);border-radius:12px;padding:16px;text-align:center;box-shadow:var(--sh)}}
.ov-gauge-lbl{{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:var(--tx2);margin-bottom:4px}}
.ov-gauge-val{{font-size:28px;font-weight:800;font-family:'JetBrains Mono',monospace}}.ov-gauge-meta{{font-size:10px;color:var(--tx2)}}
.ov-gauge.ok .ov-gauge-val{{color:var(--ok)}}.ov-gauge.bad .ov-gauge-val{{color:var(--danger)}}
/* Fleet table */
.ov-fleet th{{text-align:center;font-weight:700;color:#fff;font-size:9px;text-transform:uppercase;letter-spacing:.04em;padding:8px 6px;white-space:nowrap}}
.ov-fleet thead tr:first-child th{{background:#1F4E79}}.ov-fleet thead tr:nth-child(2) th{{background:#2E75B6}}
.ov-fleet td{{padding:6px;border-bottom:1px solid var(--s2);text-align:center;font-family:'JetBrains Mono',monospace;font-size:11px}}
.ov-fleet tr:hover td{{background:var(--rzg2)}}
.ov-fleet .equip-cell{{font-weight:700;text-align:left}}
.frb{{display:inline-block;padding:1px 6px;border-radius:4px;font-size:9px;font-weight:700;color:#fff}}
.frb-351{{background:#00843D}}.frb-352{{background:#0984e3}}.frb-353{{background:#d48806}}
.ov-fleet th.sort{{cursor:pointer;user-select:none;position:relative;padding-right:12px}}
.ov-fleet th.sort::after{{content:'⇅';position:absolute;right:1px;top:50%;transform:translateY(-50%);font-size:8px;opacity:.5}}
.ov-fleet th.sort.asc::after{{content:'↑';opacity:1}}.ov-fleet th.sort.desc::after{{content:'↓';opacity:1}}
.bar-c{{display:flex;align-items:center;gap:4px}}.bar-bg{{flex:1;height:6px;background:#eee;border-radius:3px;overflow:hidden}}.bar-f{{height:100%;border-radius:3px}}
.bar-f.high{{background:var(--danger)}}.bar-f.med{{background:#f0932b}}.bar-f.low{{background:var(--ok)}}
.meta-row{{background:var(--s2)!important}}.meta-row td{{font-weight:700;color:var(--bl);font-size:10px;border-bottom:2px solid var(--bl)!important}}
.chart-w{{position:relative;height:280px}}.chart-ww{{position:relative;height:320px}}
.ov-transport-div{{background:linear-gradient(135deg,#00843D,#004D23);color:#fff;padding:12px 18px;border-radius:12px;margin:20px 0 14px;font-size:14px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;box-shadow:0 2px 12px rgba(0,66,30,.2)}}
.hidden{{display:none}}
@media(max-width:1100px){{.frente-grid,.ov-cct-grid,.ov-gauge-row{{grid-template-columns:1fr}}.gap-grid{{grid-template-columns:1fr}}}}
@media(max-width:900px){{.cg2,.p-grid-2{{grid-template-columns:1fr}}.kg{{grid-template-columns:repeat(2,1fr)}}.ov-tcd{{grid-template-columns:1fr}}}}
@media(max-width:600px){{.kg,.frente-grid{{grid-template-columns:1fr}}.hdr{{flex-direction:column;gap:10px;text-align:center}}.tabs{{width:100%}}.tab{{flex:1;text-align:center;padding:10px 0}}}}
</style></head><body>
<div class="app">
<div class="hdr"><div class="hdr-left"><div class="hdr-logo">R</div><h1>CCT Ipaussu <span>— Painel Integrado 26/27</span></h1></div><div class="hdr-r"><span class="badge" id="spct"></span><span class="badge" id="sdt"></span></div></div>
<div class="tabs" id="master-tabs">
<button class="tab on" onclick="switchTab(0)">Visão Mensal</button>
<button class="tab" onclick="switchTab(1)">Visão por Frente</button>
<button class="tab" onclick="switchTab(2)">Perdas de Colheita</button>
<button class="tab" style="color:var(--bl);font-weight:800" onclick="switchTab(3)">📊 Overview CCT</button>
<button class="tab" onclick="switchTab(4)">✂️ Detalhe Corte</button>
<button class="tab" onclick="switchTab(5)">🚚 Detalhe Transporte</button>
</div>
<div id="tab0"><div class="filters" id="flt"></div><div id="content-tab0"></div></div>
<div id="tab1" class="hidden"><div class="filters" id="flt1"></div><div style="display:flex;gap:10px;margin-bottom:14px;align-items:center;flex-wrap:wrap"><span style="font-size:11px;font-weight:700;color:var(--tx2);text-transform:uppercase;letter-spacing:.8px">Período:</span><input type="date" id="dt1s" class="fb" style="font-family:inherit" onchange="renderProd()"><span style="font-size:11px;color:var(--tx2)">até</span><input type="date" id="dt1e" class="fb" style="font-family:inherit" onchange="renderProd()"><button class="fb" onclick="document.getElementById(\'dt1s\').value=\'\';document.getElementById(\'dt1e\').value=\'\';renderProd()">Limpar</button></div><div id="content-tab1"></div></div>
<div id="tab2" class="hidden"><div class="perdas-wrapper"><div class="p-filters" id="p-filters" style="margin-top:10px"></div><div class="p-kpi-row" id="p-kpiRow"></div><div class="p-card"><div class="p-card-header"><div class="p-card-title">Composição das Perdas (Média kg)</div></div><div class="p-card-body"><div class="p-breakdown-grid" id="p-breakdownGrid"></div></div></div><div class="p-grid-2"><div class="p-card" style="margin-bottom:0"><div class="p-card-header"><div class="p-card-title">Perdas por Frente</div></div><div class="p-card-body"><div class="p-chart-container"><canvas id="chartFrente"></canvas></div></div></div><div class="p-card" style="margin-bottom:0"><div class="p-card-header"><div class="p-card-title">Perdas por Turno</div></div><div class="p-card-body"><div class="p-chart-container"><canvas id="chartTurno"></canvas></div></div></div></div><div class="p-card"><div class="p-card-header"><div class="p-card-title">🏆 Ranking de Equipamentos por Perda</div></div><div class="p-card-body" style="padding:0"><div class="p-table-wrap"><table class="p-table"><thead><tr><th>#</th><th>Equip</th><th>Frente</th><th>Aval.</th><th>Média</th><th>Visual</th><th>Status</th><th>Touc.</th><th>Palm.</th><th>Inteira</th><th>Rebolo</th><th>Pedaço</th><th>Estil.</th><th>Lasca</th></tr></thead><tbody id="p-rankingBody"></tbody></table></div></div></div><div class="p-card"><div class="p-card-header"><div class="p-card-title">📋 Detalhamento Equip × Turno</div></div><div class="p-card-body" style="padding:0"><div class="p-table-wrap"><table class="p-table"><thead><tr><th>#</th><th>Equip.</th><th>Frente</th><th>Fazenda</th><th>Zona</th><th>Turno</th><th>Aval.</th><th>Média</th><th>Visual</th><th>Status</th></tr></thead><tbody id="p-detailBody"></tbody></table></div></div></div></div></div>
<div id="tab3" class="hidden"><div id="overviewContent"></div></div>
<div id="tab4" class="hidden"><div id="corteContent"></div></div>
<div id="tab5" class="hidden"><div id="transporteContent"></div></div>
</div>
<script>
const D={json_prod};
const MO=['Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'];
const FR={fr_list_js};
const FC={fc_obj};const FCD={fcd_obj};
let selM=['Todos'],activeTab=0,ch={{}};
const DATA_P={json_perdas};const META_P={META_PERDAS};const TOL_P={TOLERANCIA_PERDAS};
let currentFilterP='all',currentSemanaP='all',currentFazendaP='all',currentZonaP='all';
let chartFrenteInst=null,chartTurnoInst=null,perdasRendered=false;
const OV={json_overview};
let ovFrente='todas',ovRendered=false,ovSortCol='pctOfensor',ovSortDir='desc',ovCharts={{}},ovDtStart='',ovDtEnd='';
function fmt(n,d){{d=d||0;if(n==null||isNaN(n))return'—';return n.toLocaleString('pt-BR',{{minimumFractionDigits:d,maximumFractionDigits:d}})}}
Chart.register(ChartDataLabels);
let corteRendered=false,transpRendered=false;
function switchTab(t){{activeTab=t;document.querySelectorAll('.tab').forEach((b,i)=>b.classList.toggle('on',i===t));['tab0','tab1','tab2','tab3','tab4','tab5'].forEach((id,i)=>document.getElementById(id).classList.toggle('hidden',i!==t));if(t===2){{if(DATA_P&&DATA_P!=='null'){{if(!perdasRendered){{initPerdas();perdasRendered=true}}else renderPerdasAll()}}}}else if(t===3){{if(OV&&OV!=='null'){{if(!ovRendered){{renderOverviewFull();ovRendered=true}}}}}}else if(t===4){{if(OV&&OV!=='null'){{if(!corteRendered){{renderCorteDetail();corteRendered=true}}}}}}else if(t===5){{if(OV&&OV!=='null'){{if(!transpRendered){{renderTransporteDetail();transpRendered=true}}}}}}else renderProd()}}

/* ════════ PRODUÇÃO (tabs 0+1) - mesma lógica do seu código original ════════ */
function toggleMonth(m){{if(m==='Todos')selM=['Todos'];else{{if(selM.includes('Todos'))selM=[];if(selM.includes(m)){{selM=selM.filter(x=>x!==m);if(!selM.length)selM=['Todos']}}else selM.push(m)}};renderProd()}}
function fD(){{var dd=D.d.filter(d=>d.mt!=null);if(!selM.includes('Todos'))dd=dd.filter(d=>selM.includes(d.mes));var ds=document.getElementById('dt1s'),de=document.getElementById('dt1e');if(ds&&ds.value)dd=dd.filter(d=>d.dia>=ds.value);if(de&&de.value)dd=dd.filter(d=>d.dia<=de.value);return dd}}
function rF(){{document.getElementById('flt').innerHTML=['Todos'].concat(MO).map(m=>'<button class="fb '+(selM.includes(m)?'on':'')+'" onclick="toggleMonth(\\''+m+'\\')">'+m+'</button>').join('')}}
function renderProd(){{rF();if(activeTab===0)renderTab0();else if(activeTab===1)renderTab1()}}
function renderTab0(){{var dd=fD(),el=document.getElementById('content-tab0');var tM=0,tP=0;dd.forEach(d=>{{tM+=d.mt||0;tP+=d.mp||0}});var f1=FR[0];var fa=dd.filter(d=>d['a'+f1]);var fi=dd.filter(d=>d['i'+f1]);var fr2=dd.filter(d=>d['k'+f1]);var fd=dd.filter(d=>d['d'+f1]);var aA=0,aI=0,aR=0,aD2=0;if(fa.length)fa.forEach(d=>aA+=d['a'+f1]/fa.length);if(fi.length)fi.forEach(d=>aI+=d['i'+f1]/fi.length);if(fr2.length)fr2.forEach(d=>aR+=d['k'+f1]/fr2.length);if(fd.length)fd.forEach(d=>aD2+=d['d'+f1]/fd.length);var mx=tM>0?(tP/tM*100):0;var mT=0,rT=0;MO.forEach(m=>{{if(selM.includes('Todos')||selM.includes(m)){{mT+=D.m.meta_ipa[m]||0;rT+=D.m.real_u[m]||0}}}});if(selM.includes('Todos')){{mT=D.m.meta_ipa.Total;rT=D.m.real_u.Total}};var pM=mT>0?(rT/mT*100):0;document.getElementById('spct').textContent='Safra: '+fmt(pM,1)+'%';var ld=dd[dd.length-1];document.getElementById('sdt').textContent=ld?'Último: '+ld.dia:'';var h='<div class="kg"><div class="k cg"><div class="kl">Moagem Total</div><div class="kv">'+fmt(tM)+'</div><div class="ks">ton</div><div class="kp '+(pM>=100?'pu':'pd')+'">'+fmt(pM,1)+'% da meta</div></div><div class="k cb"><div class="kl">Moagem Própria</div><div class="kv">'+fmt(tP)+'</div><div class="ks">Mix '+fmt(mx,1)+'%</div></div><div class="k cg"><div class="kl">ATR Médio</div><div class="kv">'+fmt(aA,1)+'</div></div><div class="k ca"><div class="kl">Impureza</div><div class="kv">'+fmt(aI,1)+'%</div></div><div class="k cc"><div class="kl">Raio</div><div class="kv">'+fmt(aR,1)+' km</div></div><div class="k cp"><div class="kl">Dens Carga</div><div class="kv">'+fmt(aD2,1)+'</div></div></div>';h+='<div class="cg2"><div class="cc full"><div class="ct"><div class="dot" style="background:var(--rz)"></div>Moagem Diária vs Plano</div><div class="cw tall"><canvas id="mc1"></canvas></div></div><div class="cc"><div class="ct"><div class="dot" style="background:var(--bl)"></div>Meta vs Realizado</div><div class="cw"><canvas id="mc2"></canvas></div></div><div class="cc"><div class="ct"><div class="dot" style="background:var(--cy)"></div>ATR Médio</div><div class="cw"><canvas id="mc3"></canvas></div></div><div class="cc"><div class="ct"><div class="dot" style="background:var(--am)"></div>Impureza Média</div><div class="cw"><canvas id="mc4"></canvas></div></div><div class="cc"><div class="ct"><div class="dot" style="background:var(--pr)"></div>Mix e Raio</div><div class="cw"><canvas id="mc5"></canvas></div></div></div>';h+='<div class="tc"><div class="ct"><div class="dot" style="background:var(--rz)"></div>Resumo Mensal</div><table><thead><tr><th style="text-align:left">Mês</th><th>Meta IPA</th><th>Realizado</th><th>%</th><th>Meta Próp</th><th>Real Próp</th><th>%</th></tr></thead><tbody>';MO.forEach(m=>{{var mi=D.m.meta_ipa[m]||0,ri=D.m.real_u[m]||0,pi=mi>0?ri/mi*100:0;var mp2=D.m.meta_p[m]||0,rp=D.m.real_p[m]||0,pp=mp2>0?rp/mp2*100:0;h+='<tr><td style="font-weight:700;text-align:left">'+m+'</td><td>'+fmt(mi)+'</td><td>'+fmt(ri)+'</td><td class="'+(pi>=100?'vp':'vn')+'">'+fmt(pi,1)+'%</td><td>'+fmt(mp2)+'</td><td>'+fmt(rp)+'</td><td class="'+(pp>=100?'vp':'vn')+'">'+fmt(pp,1)+'%</td></tr>'}});var ti=D.m.meta_ipa.Total||0,tri=D.m.real_u.Total||0,tp=D.m.meta_p.Total||0,trp=D.m.real_p.Total||0;h+='<tr style="font-weight:700;border-top:2px solid var(--rz)"><td style="text-align:left">TOTAL</td><td>'+fmt(ti)+'</td><td>'+fmt(tri)+'</td><td class="'+(tri/ti>=1?'vp':'vn')+'">'+fmt(tri/ti*100,1)+'%</td><td>'+fmt(tp)+'</td><td>'+fmt(trp)+'</td><td class="'+(trp/tp>=1?'vp':'vn')+'">'+fmt(trp/tp*100,1)+'%</td></tr></tbody></table></div>';el.innerHTML=h;renderMC(dd)}}
function renderMC(dd){{Object.keys(ch).forEach(k=>{{if(k[0]==='m')ch[k].destroy()}});var gc='#e8ede9',tc2='#5a7265';Chart.defaults.color=tc2;Chart.defaults.font.family='Plus Jakarta Sans';var x1=document.getElementById('mc1');if(x1)ch.mc1=new Chart(x1,{{type:'bar',data:{{labels:dd.map(d=>d.dia.slice(5)),datasets:[{{label:'Moagem',data:dd.map(d=>d.mt),backgroundColor:'rgba(0,132,61,.2)',borderColor:'#00843D',borderWidth:1,borderRadius:3,order:2}},{{label:'Plano',data:dd.map(()=>2448.5),type:'line',borderColor:'#e17055',borderWidth:2,borderDash:[6,3],pointRadius:0,order:1,fill:false}}]}},options:{{layout:{{padding:{{top:20}}}},responsive:true,maintainAspectRatio:false,plugins:{{datalabels:{{display:false}},legend:{{labels:{{color:tc2}}}}}},scales:{{x:{{ticks:{{color:tc2,font:{{size:8}},maxRotation:45,autoSkip:true,maxTicksLimit:30}},grid:{{color:gc}}}},y:{{ticks:{{color:tc2,callback:v=>fmt(v)}},grid:{{color:gc}}}}}}}}}});var ms=MO.filter(m=>(D.m.meta_ipa[m]||0)>0||(D.m.real_u[m]||0)>0);var x2=document.getElementById('mc2');if(x2)ch.mc2=new Chart(x2,{{type:'bar',data:{{labels:ms,datasets:[{{label:'Meta',data:ms.map(m=>D.m.meta_ipa[m]||0),backgroundColor:'rgba(225,112,85,.25)',borderColor:'#e17055',borderWidth:1.5,borderRadius:4}},{{label:'Real',data:ms.map(m=>D.m.real_u[m]||0),backgroundColor:'rgba(0,132,61,.2)',borderColor:'#00843D',borderWidth:1.5,borderRadius:4}}]}},options:{{layout:{{padding:{{top:20}}}},responsive:true,maintainAspectRatio:false,plugins:{{datalabels:{{display:false}}}},scales:{{x:{{grid:{{color:gc}}}},y:{{ticks:{{callback:v=>(v/1000)+'k'}},grid:{{color:gc}}}}}}}}}});var f1=FR[0];var avgA=MO.map(m=>{{var d=D.d.filter(r=>r.mes===m&&r['a'+f1]);if(!d.length)return null;var s=0;d.forEach(r=>s+=r['a'+f1]);return s/d.length}});var x3=document.getElementById('mc3');if(x3)ch.mc3=new Chart(x3,{{type:'line',data:{{labels:MO,datasets:[{{data:avgA,borderColor:'#00b894',backgroundColor:'rgba(0,184,148,.08)',fill:true,tension:.3,pointBackgroundColor:'#00b894',pointRadius:5,borderWidth:2.5}}]}},options:{{layout:{{padding:{{top:20}}}},responsive:true,maintainAspectRatio:false,plugins:{{datalabels:{{display:false}},legend:{{display:false}}}},scales:{{x:{{grid:{{color:gc}}}},y:{{grid:{{color:gc}}}}}}}}}});var avgI=MO.map(m=>{{var d=D.d.filter(r=>r.mes===m&&r['i'+f1]);if(!d.length)return null;var s=0;d.forEach(r=>s+=r['i'+f1]);return s/d.length}});var x4=document.getElementById('mc4');if(x4)ch.mc4=new Chart(x4,{{type:'bar',data:{{labels:MO,datasets:[{{data:avgI,backgroundColor:avgI.map(v=>v>8?'rgba(214,48,49,.3)':'rgba(0,132,61,.2)'),borderColor:avgI.map(v=>v>8?'#d63031':'#00843D'),borderWidth:1.5,borderRadius:4}}]}},options:{{layout:{{padding:{{top:20}}}},responsive:true,maintainAspectRatio:false,plugins:{{datalabels:{{display:false}},legend:{{display:false}}}},scales:{{x:{{grid:{{color:gc}}}},y:{{grid:{{color:gc}}}}}}}}}});var avgMx=MO.map(m=>{{var d=D.d.filter(r=>r.mes===m&&r.mix_p);if(!d.length)return null;var s=0;d.forEach(r=>s+=r.mix_p);return s/d.length}});var avgR=MO.map(m=>{{var d=D.d.filter(r=>r.mes===m&&r['k'+f1]);if(!d.length)return null;var s=0;d.forEach(r=>s+=r['k'+f1]);return s/d.length}});var x5=document.getElementById('mc5');if(x5)ch.mc5=new Chart(x5,{{type:'bar',data:{{labels:MO,datasets:[{{label:'Mix%',data:avgMx,backgroundColor:'rgba(108,92,231,.2)',borderColor:'#6c5ce7',borderWidth:1.5,borderRadius:4,yAxisID:'y'}},{{label:'Raio',data:avgR,type:'line',borderColor:'#e17055',pointBackgroundColor:'#e17055',pointRadius:4,borderWidth:2,yAxisID:'y1'}}]}},options:{{layout:{{padding:{{top:20}}}},responsive:true,maintainAspectRatio:false,plugins:{{datalabels:{{display:false}}}},scales:{{x:{{grid:{{color:gc}}}},y:{{position:'left',grid:{{color:gc}}}},y1:{{position:'right',grid:{{display:false}}}}}}}}}})}}
function calcFr(dd){{return FR.map(f=>{{var rows=dd.filter(d=>d['r'+f]!=null);var tR=0,wA=0,wI=0,wK=0,wD=0,tA=0,tI=0,tK=0,tD=0;rows.forEach(d=>{{tR+=d['r'+f]||0;if(d['a'+f]){{wA+=d['a'+f]*d['r'+f];tA+=d['r'+f]}};if(d['i'+f]){{wI+=d['i'+f]*d['r'+f];tI+=d['r'+f]}};if(d['k'+f]){{wK+=d['k'+f]*d['r'+f];tK+=d['r'+f]}};if(d['d'+f]){{wD+=d['d'+f]*d['r'+f];tD+=d['r'+f]}}}});return{{f:f,days:rows.length,real:tR,atr:tA>0?wA/tA:0,imp:tI>0?wI/tI:0,raio:tK>0?wK/tK:0,dc:tD>0?wD/tD:0,mix:0}}}})}}
function calcMF(){{return MO.map(m=>{{var dd=D.d.filter(d=>d.mes===m&&d.mt!=null);var fr3=calcFr(dd);var tot=0;fr3.forEach(x=>tot+=x.real);fr3.forEach(x=>x.mix=tot>0?x.real/tot*100:0);return{{mes:m,frentes:fr3,totalMoagem:tot}}}}).filter(x=>x.totalMoagem>0)}}
function renderTab1(){{var dd=fD();var fr3=calcFr(dd);var tot=0;fr3.forEach(x=>tot+=x.real);fr3.forEach(x=>x.mix=tot>0?x.real/tot*100:0);var h='<div class="section-title"><div class="bar"></div>Resumo por Frente</div><div class="frente-grid">';fr3.forEach(x=>{{var ic=x.imp>8?'var(--red)':x.imp>6?'var(--am)':'var(--rz)';h+='<div class="fc" style="--fc:'+FC[x.f]+'"><div style="position:absolute;top:0;left:0;width:100%;height:3px;background:'+FC[x.f]+'"></div><div class="fc-name"><div class="dot" style="background:'+FC[x.f]+'"></div>Frente '+x.f+'</div><div class="fc-row"><span class="label">Moagem</span><span class="val">'+fmt(x.real)+' t</span></div><div class="fc-row"><span class="label">Mix</span><span class="val">'+fmt(x.mix,1)+'%</span></div><div class="fc-row"><span class="label">ATR</span><span class="val">'+fmt(x.atr,1)+'</span></div><div class="fc-row"><span class="label">Impureza</span><span class="val" style="color:'+ic+'">'+fmt(x.imp,1)+'%</span></div><div class="fc-row"><span class="label">Raio</span><span class="val">'+fmt(x.raio,1)+' km</span></div><div class="fc-row"><span class="label">Dens.Carga</span><span class="val">'+fmt(x.dc,1)+' t</span></div><div class="fc-row"><span class="label">Dias</span><span class="val">'+x.days+'</span></div></div>'}});h+='</div>';h+='<div class="section-title"><div class="bar"></div>Análise de Gaps</div><div class="gap-grid">';var maxATR=0;fr3.forEach(x=>{{if(x.atr>maxATR)maxATR=x.atr}});var best=fr3[0];fr3.forEach(x=>{{if(x.atr>best.atr)best=x}});var sA=fr3.slice().sort((a,b)=>b.atr-a.atr);h+='<div class="gap-card"><h3>ATR por Frente</h3>';sA.forEach(x=>{{var pct=maxATR>0?(x.atr/maxATR*100):0;var diff=x.atr-best.atr;h+='<div class="gap-bar-row"><div class="gap-label">'+x.f+'</div><div class="gap-bar-wrap"><div class="gap-bar" style="width:'+pct+'%;background:'+FC[x.f]+'">'+fmt(x.atr,1)+'</div></div>';var dv=diff===0?'TOP':'<span class="gap-diff '+(diff>=0?'pos':'neg')+'">'+(diff>0?'+':'')+fmt(diff,1)+'</span>';h+='<div class="gap-val">'+dv+'</div></div>'}});h+='</div>';var sI=fr3.slice().sort((a,b)=>a.imp-b.imp);var maxImp=0;fr3.forEach(x=>{{if(x.imp>maxImp)maxImp=x.imp}});var bestImp=sI[0];h+='<div class="gap-card"><h3>Impureza por Frente (%)</h3>';sI.forEach(x=>{{var pct=maxImp>0?(x.imp/maxImp*100):0;var diff=x.imp-bestImp.imp;var bc=x.imp>10?'#d63031':x.imp>8?'#e17055':FC[x.f];h+='<div class="gap-bar-row"><div class="gap-label">'+x.f+'</div><div class="gap-bar-wrap"><div class="gap-bar" style="width:'+pct+'%;background:'+bc+'">'+fmt(x.imp,1)+'%</div></div>';var dv2=diff===0?'MELHOR':'<span class="gap-diff neg">+'+fmt(diff,1)+'pp</span>';h+='<div class="gap-val">'+dv2+'</div></div>'}});h+='</div>';var sM2=fr3.slice().sort((a,b)=>b.mix-a.mix);h+='<div class="gap-card"><h3>Participação Moagem</h3>';sM2.forEach(x=>{{h+='<div class="gap-bar-row"><div class="gap-label">'+x.f+'</div><div class="gap-bar-wrap"><div class="gap-bar" style="width:'+x.mix+'%;background:'+FC[x.f]+'">'+fmt(x.mix,1)+'%</div></div><div class="gap-val">'+fmt(x.real)+' t</div></div>'}});h+='</div></div>';h+='<div class="cg2"><div class="cc"><div class="ct"><div class="dot" style="background:var(--rz)"></div>ATR Mensal por Frente</div><div class="cw"><canvas id="fc1"></canvas></div></div><div class="cc"><div class="ct"><div class="dot" style="background:var(--red)"></div>Impureza Mensal</div><div class="cw"><canvas id="fc2"></canvas></div></div><div class="cc"><div class="ct"><div class="dot" style="background:var(--pr)"></div>Mix por Frente</div><div class="cw"><canvas id="fc3"></canvas></div></div><div class="cc"><div class="ct"><div class="dot" style="background:var(--am)"></div>Moagem por Frente</div><div class="cw"><canvas id="fc4"></canvas></div></div></div>';document.getElementById('content-tab1').innerHTML=h;renderFC()}}
function renderFC(){{Object.keys(ch).forEach(k=>{{if(k[0]==='f')ch[k].destroy()}});var mf=calcMF();var labels=mf.map(x=>x.mes);var gc='#e8ede9';function gd(prop){{return FR.map(f=>({{label:'F'+f,data:mf.map(x=>{{var v=null;x.frentes.forEach(y=>{{if(y.f===f&&y[prop]>0)v=y[prop]}});return v}}),borderColor:FC[f],backgroundColor:FCD[f],tension:.3,pointRadius:4,pointBackgroundColor:FC[f],borderWidth:2.5,spanGaps:true}}))}}var dlO={{color:'#1a2e22',font:{{weight:'bold',size:9,family:'JetBrains Mono'}},backgroundColor:'rgba(255,255,255,0.75)',borderRadius:3,padding:{{top:2,bottom:2,left:4,right:4}},offset:2,display:ctx=>ctx.dataIndex===ctx.dataset.data.length-1,formatter:v=>v?fmt(v,1):''}};var opts={{layout:{{padding:{{top:20,right:30}}}},responsive:true,maintainAspectRatio:false,plugins:{{datalabels:dlO,legend:{{labels:{{color:'#5a7265'}}}}}},scales:{{x:{{grid:{{color:gc}}}},y:{{grid:{{color:gc}}}}}}}};var c1=document.getElementById('fc1');if(c1)ch.fc1=new Chart(c1,{{type:'line',data:{{labels:labels,datasets:gd('atr')}},options:opts}});var c2=document.getElementById('fc2');if(c2)ch.fc2=new Chart(c2,{{type:'line',data:{{labels:labels,datasets:gd('imp')}},options:opts}});var c3=document.getElementById('fc3');if(c3){{var ds3=FR.map(f=>({{label:'F'+f,data:mf.map(x=>{{var v=0;x.frentes.forEach(y=>{{if(y.f===f)v=y.mix}});return v}}),backgroundColor:FCD[f],borderColor:FC[f],borderWidth:1.5,borderRadius:3}}));ch.fc3=new Chart(c3,{{type:'bar',data:{{labels:labels,datasets:ds3}},options:{{layout:{{padding:{{top:20}}}},responsive:true,maintainAspectRatio:false,plugins:{{datalabels:{{display:false}},legend:{{labels:{{color:'#5a7265'}}}}}},scales:{{x:{{stacked:true,grid:{{color:gc}}}},y:{{stacked:true,grid:{{color:gc}}}}}}}}}})}}var c4=document.getElementById('fc4');if(c4){{var ds4=FR.map(f=>({{label:'F'+f,data:mf.map(x=>{{var v=0;x.frentes.forEach(y=>{{if(y.f===f)v=y.real}});return v}}),backgroundColor:FCD[f],borderColor:FC[f],borderWidth:1.5,borderRadius:3}}));ch.fc4=new Chart(c4,{{type:'bar',data:{{labels:labels,datasets:ds4}},options:{{layout:{{padding:{{top:20}}}},responsive:true,maintainAspectRatio:false,plugins:{{datalabels:{{display:false}},legend:{{labels:{{color:'#5a7265'}}}}}},scales:{{x:{{grid:{{color:gc}}}},y:{{ticks:{{callback:v=>(v/1000)+'k'}},grid:{{color:gc}}}}}}}}}})}}}}

/* ════════ PERDAS ════════ */
function getStatusP(val){{if(val>META_P)return'red';if(val>=META_P-TOL_P)return'yellow';return'green'}}
function getStatusLabelP(val){{const s=getStatusP(val);if(s==='red')return'⚠️ Acima';if(s==='yellow')return'🟡 Atenção';return'✅ Dentro'}}
function statusPillP(val){{const s=getStatusP(val);return`<span class="p-status-pill p-pill-${{s}}">${{getStatusLabelP(val)}}</span>`}}
function rankClassP(i){{if(i<=2)return'p-rank-1';if(i<=4)return'p-rank-2';if(i<=6)return'p-rank-3';return'p-rank-default'}}
function groupByP(data,keys){{const map={{}};const fields=['agr_touceira','agr_palmito','cana_inteira','rebolo','pedaco','estilhaco','lasca'];data.forEach(r=>{{const key=keys.map(k=>r[k]).join('|');if(!map[key]){{map[key]={{count:0,_sum:0,_cols:{{}}}};keys.forEach(k=>map[key][k]=r[k]);fields.forEach(c=>map[key]._cols[c]=0)}};map[key].count+=r.count;map[key]._sum+=(r.media*r.count);fields.forEach(c=>{{map[key]._cols[c]+=(r[c]||0)*r.count}})}});return Object.values(map).map(g=>{{const out={{count:g.count,media:g._sum/g.count}};keys.forEach(k=>out[k]=g[k]);fields.forEach(c=>{{out[c]=g._cols[c]/g.count}});return out}})}}
function getFilteredRawP(){{return DATA_P.raw.filter(r=>(currentFilterP==='all'||String(r.frente)===currentFilterP)&&(currentSemanaP==='all'||r.semana===currentSemanaP)&&(currentFazendaP==='all'||String(r.fazenda)===currentFazendaP)&&(currentZonaP==='all'||String(r.zona)===currentZonaP))}}
function initPerdas(){{if(!DATA_P||!DATA_P.raw){{document.getElementById('tab2').innerHTML='<div style="padding:40px;text-align:center;color:var(--tx2)">Aba Base colheita não encontrada.</div>';return}};const fEl=document.getElementById('p-filters');let h=`<div class="p-filter-label">Frente</div><button class="p-filter-btn active" data-frente="all" onclick="setFilterP('all')">Todas</button>`;DATA_P.frentes.forEach(f=>h+=`<button class="p-filter-btn" data-frente="${{f}}" onclick="setFilterP('${{f}}')">Frente ${{f}}</button>`);h+=`<div class="p-filter-label" style="margin-left:16px">Semana</div><select class="p-filter-btn active" onchange="setSemanaP(this.value)"><option value="all">Todas</option>`;DATA_P.semanas.forEach(s=>h+=`<option value="${{s}}">${{s}}</option>`);h+=`</select>`;h+=`<div class="p-filter-label" style="margin-left:16px">Fazenda</div><select class="p-filter-btn active" onchange="setFazendaP(this.value)"><option value="all">Todas</option>`;DATA_P.fazendas.forEach(fz=>{{if(fz!=='S/ Fazenda')h+=`<option value="${{fz}}">${{fz}}</option>`}});h+=`</select>`;h+=`<div class="p-filter-label" style="margin-left:16px">Zona</div><select class="p-filter-btn active" onchange="setZonaP(this.value)"><option value="all">Todas</option>`;DATA_P.zonas.forEach(z=>{{if(z!=='S/ Zona')h+=`<option value="${{z}}">${{z}}</option>`}});h+=`</select>`;fEl.innerHTML=h;renderPerdasAll()}}
function setFilterP(f){{currentFilterP=f;document.querySelectorAll('.p-filter-btn[data-frente]').forEach(b=>b.classList.toggle('active',b.dataset.frente===String(f)));renderPerdasAll()}}
function setSemanaP(s){{currentSemanaP=s;renderPerdasAll()}}
function setFazendaP(fz){{currentFazendaP=fz;renderPerdasAll()}}
function setZonaP(z){{currentZonaP=z;renderPerdasAll()}}
function renderPerdasAll(){{const fr=getFilteredRawP();if(!fr.length){{document.getElementById('p-kpiRow').innerHTML='<div style="color:var(--tx2);padding:20px">Sem dados para este filtro.</div>';document.getElementById('p-breakdownGrid').innerHTML='';document.getElementById('p-rankingBody').innerHTML='';document.getElementById('p-detailBody').innerHTML='';if(chartFrenteInst)chartFrenteInst.destroy();if(chartTurnoInst)chartTurnoInst.destroy();return}};const agg=groupByP(fr,['dummy'])[0];const s=getStatusP(agg.media);const eqFr=groupByP(fr,['equip','frente']);const above=eqFr.filter(r=>r.media>META_P).length;const attn=eqFr.filter(r=>r.media>=META_P-TOL_P&&r.media<=META_P).length;document.getElementById('p-kpiRow').innerHTML=`<div class="p-kpi-card status-${{s}}"><div class="p-kpi-label">Média Geral</div><div class="p-kpi-value p-text-${{s}}">${{fmt(agg.media,2)}} <span class="p-kpi-unit">kg</span></div><div class="p-kpi-status">${{statusPillP(agg.media)}}</div></div><div class="p-kpi-card status-green"><div class="p-kpi-label">Avaliações</div><div class="p-kpi-value">${{agg.count.toLocaleString('pt-BR')}}</div></div><div class="p-kpi-card ${{above>0?'status-red':'status-green'}}"><div class="p-kpi-label">Acima Meta</div><div class="p-kpi-value ${{above>0?'p-text-red':'p-text-green'}}">${{above}}</div></div><div class="p-kpi-card ${{attn>0?'status-yellow':'status-green'}}"><div class="p-kpi-label">Em Atenção</div><div class="p-kpi-value ${{attn>0?'p-text-yellow':'p-text-green'}}">${{attn}}</div></div>`;const items=[{{key:'pedaco',label:'Pedaço'}},{{key:'agr_touceira',label:'Agr.Touceira'}},{{key:'lasca',label:'Lasca'}},{{key:'cana_inteira',label:'Cana Inteira'}},{{key:'rebolo',label:'Rebolo'}},{{key:'estilhaco',label:'Estilhaço'}},{{key:'agr_palmito',label:'Agr.Palmito'}}].sort((a,b)=>(agg[b.key]||0)-(agg[a.key]||0));const mx=Math.max(...items.map(i=>agg[i.key]||0));document.getElementById('p-breakdownGrid').innerHTML=items.map(i=>{{const v=agg[i.key]||0;const p=mx>0?(v/mx*100):0;return`<div class="p-breakdown-item"><div class="p-breakdown-label">${{i.label}}</div><div class="p-breakdown-value">${{fmt(v,2)}}</div><div class="p-breakdown-bar"><div class="p-breakdown-bar-fill" style="width:${{p}}%"></div></div></div>`}}).join('');const fD2=groupByP(fr,['frente']).sort((a,b)=>b.media-a.media);chartFrenteInst=makeBarP('chartFrente',fD2.map(r=>'Fr '+r.frente),fD2.map(r=>r.media),chartFrenteInst);const tD2=groupByP(fr,['turno']).sort((a,b)=>a.turno.localeCompare(b.turno));chartTurnoInst=makeBarP('chartTurno',tD2.map(r=>'T '+r.turno),tD2.map(r=>r.media),chartTurnoInst);const mxM=Math.max(...eqFr.map(r=>r.media),META_P);document.getElementById('p-rankingBody').innerHTML=eqFr.sort((a,b)=>b.media-a.media).map((r,i)=>{{const s2=getStatusP(r.media);const bp=(r.media/(mxM*1.1))*100;return`<tr><td><span class="p-rank-num ${{rankClassP(i+1)}}">${{i+1}}</span></td><td class="p-mono" style="font-weight:600">${{r.equip}}</td><td>${{r.frente}}</td><td class="p-text-muted">${{r.count}}</td><td class="p-mono p-text-${{s2}}" style="font-weight:600">${{fmt(r.media,2)}}</td><td style="min-width:120px"><div class="p-bar-cell"><div class="p-bar-track"><div class="p-bar-fill" style="width:${{bp}}%;background:${{s2==='red'?'var(--p-red)':s2==='yellow'?'var(--p-yellow)':'var(--p-green)'}}"></div></div></div></td><td>${{statusPillP(r.media)}}</td><td class="p-mono p-text-muted">${{fmt(r.agr_touceira,2)}}</td><td class="p-mono p-text-muted">${{fmt(r.agr_palmito,2)}}</td><td class="p-mono p-text-muted">${{fmt(r.cana_inteira,2)}}</td><td class="p-mono p-text-muted">${{fmt(r.rebolo,2)}}</td><td class="p-mono p-text-muted">${{fmt(r.pedaco,2)}}</td><td class="p-mono p-text-muted">${{fmt(r.estilhaco,2)}}</td><td class="p-mono p-text-muted">${{fmt(r.lasca,2)}}</td></tr>`}}).join('');const dD2=groupByP(fr,['equip','frente','fazenda','zona','turno']).sort((a,b)=>b.media-a.media);document.getElementById('p-detailBody').innerHTML=dD2.map((r,i)=>{{const s2=getStatusP(r.media);const bp=(r.media/(mxM*1.1))*100;return`<tr><td><span class="p-rank-num ${{rankClassP(i+1)}}">${{i+1}}</span></td><td class="p-mono" style="font-weight:600">${{r.equip}}</td><td>${{r.frente}}</td><td style="font-size:11px;color:var(--tx2)">${{r.fazenda}}</td><td style="font-size:11px;color:var(--tx2)">${{r.zona}}</td><td style="font-weight:600">T ${{r.turno}}</td><td class="p-text-muted">${{r.count}}</td><td class="p-mono p-text-${{s2}}" style="font-weight:600">${{fmt(r.media,2)}}</td><td style="min-width:100px"><div class="p-bar-cell"><div class="p-bar-track"><div class="p-bar-fill" style="width:${{bp}}%;background:${{s2==='red'?'var(--p-red)':s2==='yellow'?'var(--p-yellow)':'var(--p-green)'}}"></div></div></div></td><td>${{statusPillP(r.media)}}</td></tr>`}}).join('')}}
function makeBarP(id,labels,values,inst){{if(inst)inst.destroy();return new Chart(document.getElementById(id).getContext('2d'),{{type:'bar',data:{{labels,datasets:[{{data:values,backgroundColor:values.map(v=>{{const s=getStatusP(v);return s==='red'?'rgba(214,48,49,.7)':s==='yellow'?'rgba(225,112,85,.7)':'rgba(0,132,61,.7)'}}),borderColor:values.map(v=>{{const s=getStatusP(v);return s==='red'?'#d63031':s==='yellow'?'#e17055':'#00843D'}}),borderWidth:1,borderRadius:4}}]}},options:{{responsive:true,maintainAspectRatio:false,layout:{{padding:{{top:30}}}},plugins:{{legend:{{display:false}},datalabels:{{display:true,color:'#1a2e22',anchor:'end',align:'top',font:{{family:'JetBrains Mono',size:11,weight:'bold'}},formatter:v=>fmt(v,2)}}}},scales:{{y:{{beginAtZero:true,grid:{{color:'#e8ede9'}}}},x:{{grid:{{display:false}}}}}}}},plugins:[{{afterDraw(c){{const ctx2=c.ctx;const y=c.scales.y.getPixelForValue(META_P);ctx2.save();ctx2.strokeStyle='#d63031';ctx2.lineWidth=1.5;ctx2.setLineDash([6,4]);ctx2.beginPath();ctx2.moveTo(c.scales.x.left,y);ctx2.lineTo(c.scales.x.right,y);ctx2.stroke();ctx2.fillStyle='#d63031';ctx2.font='10px JetBrains Mono';ctx2.fillText('META '+fmt(META_P,2),c.scales.x.right-60,y-5);ctx2.restore()}}}}]}})}}

/* ════════ OVERVIEW CCT (Tab 3) ════════ */
function killOvChart(id){{if(ovCharts[id]){{ovCharts[id].destroy();delete ovCharts[id]}}}}
function mkOvChart(id,cfg){{killOvChart(id);const el=document.getElementById(id);if(!el)return;ovCharts[id]=new Chart(el.getContext('2d'),cfg)}}
function getOvFleet(f){{
  if(!OV||!OV.raw_bordo)return f==='todas'?[...(OV.fleet||[])]:((OV.fleet||[]).filter(r=>r.frente===f));
  const trips=(OV.raw_trips||[]).filter(t=>(f==='todas'||t.fr===f)&&(!ovDtStart||t.dt>=ovDtStart)&&(!ovDtEnd||t.dt<=ovDtEnd));
  const bordo=(OV.raw_bordo||[]).filter(b=>(!ovDtStart||b.dt>=ovDtStart)&&(!ovDtEnd||b.dt<=ovDtEnd));
  const eqFr={{}};trips.forEach(t=>{{if(!eqFr[t.eq])eqFr[t.eq]=t.fr}});
  const eqs=new Set([...trips.map(t=>t.eq),...bordo.map(b=>b.eq)]);
  const out=[];
  eqs.forEach(eq=>{{
    const fr2=eqFr[eq]||(OV.fleet||[]).find(x=>x.equip===eq)?.frente||'';
    if(f!=='todas'&&fr2!==f)return;
    const bt=trips.filter(t=>t.eq===eq);const bb=bordo.filter(b=>b.eq===eq);
    const sumB=(k)=>bb.reduce((s,r)=>s+(r[k]||0),0);
    const vpv=sumB('vpv'),hov=sumB('hov'),vpc=sumB('vpc'),hoc=sumB('hoc');
    const tot=sumB('tot');if(tot<=0)return;
    out.push({{
      equip:eq,frente:fr2,
      velVazio:hov>0?Math.round(vpv/hov*100)/100:0,
      velCarreg:hoc>0?Math.round(vpc/hoc*100)/100:0,
      t1:bt.length?Math.round(bt.reduce((s,t)=>s+t.t1,0)/bt.length*10)/10:0,
      t2:bt.length?Math.round(bt.reduce((s,t)=>s+t.t2,0)/bt.length*10)/10:0,
      raio:bt.filter(t=>t.raio>0).length?Math.round(bt.filter(t=>t.raio>0).reduce((s,t)=>s+t.raio,0)/bt.filter(t=>t.raio>0).length*100)/100:0,
      densidade:bt.filter(t=>t.dens>0).length?Math.round(bt.filter(t=>t.dens>0).reduce((s,t)=>s+t.dens,0)/bt.filter(t=>t.dens>0).length*100)/100:0,
      manut:Math.round(sumB('man')*100)/100,agTransb:Math.round(sumB('ag')*100)/100,
      semApont:Math.round(sumB('sa')*100)/100,patio:Math.round(sumB('pat')*100)/100,
      totalH:Math.round(tot*100)/100,
      pctOfensor:Math.round((sumB('man')+sumB('ag')+sumB('pat'))/tot*1000)/10,
    }});
  }});
  return out;
}}

function renderOverviewFull(){{
  if(!OV||!OV.kpis){{document.getElementById('overviewContent').innerHTML='<div style="padding:40px;text-align:center;color:var(--tx2)">Arquivo Metricas_cct_dashboard.xlsx não encontrado ou sem abas de apoio.</div>';return}}
  const minDt=OV.dates&&OV.dates.length?OV.dates[0]:'';
  const maxDt=OV.dates&&OV.dates.length?OV.dates[OV.dates.length-1]:'';
  if(!document.getElementById('ovDtS')||!document.getElementById('ovDtS').value){{ovDtStart=minDt;ovDtEnd=maxDt}}
  let fh='<div class="p-filters" style="margin-bottom:16px"><div class="p-filter-label">Frente</div>';
  OV.frentes.forEach(f=>{{const act=f===ovFrente?'active':'';const lbl=f==='todas'?'Todas':'Frente '+f;fh+=`<button class="p-filter-btn ${{act}}" onclick="setOvFrente('${{f}}')">${{lbl}}</button>`}});
  fh+=`<div class="p-filter-label" style="margin-left:16px">Período</div><input type="date" id="ovDtS" class="p-filter-btn" value="${{ovDtStart}}" min="${{minDt}}" max="${{maxDt}}" onchange="ovDtStart=this.value;renderOverviewFull();ovRendered=true"><span style="font-size:11px;color:var(--tx2);padding:0 4px">até</span><input type="date" id="ovDtE" class="p-filter-btn" value="${{ovDtEnd}}" min="${{minDt}}" max="${{maxDt}}" onchange="ovDtEnd=this.value;renderOverviewFull();ovRendered=true">`;
  fh+='</div>';

  const d=OV.kpis[ovFrente]||OV.kpis['todas'];
  const M=OV.metas;
  const fl=getOvFleet(ovFrente);
  const tcdOk=d.tcd.real>=d.tcd.meta;
  const pctTcd=d.tcd.meta>0?((d.tcd.real/d.tcd.meta)*100).toFixed(1):'0';

  function kpiCard(item){{
    const ok=item.higher_is_better?item.real>=item.meta:item.real<=item.meta;
    const click=item.tab!==undefined?` style="cursor:pointer" onclick="switchTab(${{item.tab}})"`:''
    return`<div class="ov-kpi ${{ok?'ok':'bad'}}"${{click}}><div><div class="ov-kpi-nm">${{item.name}}</div><div class="ov-kpi-vl">${{fmt(item.real)}}</div><div class="ov-kpi-mt">Meta: <b>${{fmt(item.meta)}} ${{item.unit}}</b></div></div><div class="ov-kpi-ic">${{ok?'✓':'!'}}</div></div>`;
  }}

  let h = fh;
  h+=`<div class="ov-tcd ${{tcdOk?'ok':'bad'}}"><div><div class="ov-tcd-lbl">Indicador Supremo</div><div class="ov-tcd-ttl">TCD — Toneladas de Cana por Dia</div><div class="ov-tcd-vals"><div class="ov-tcd-val"><span>Realizado</span><strong>${{fmt(d.tcd.real,2)}} ton</strong></div><div class="ov-tcd-val meta"><span>Meta</span><strong>${{fmt(d.tcd.meta)}} ton</strong></div><div class="ov-tcd-val"><span>Atingimento</span><strong>${{pctTcd}}%</strong></div></div></div><div class="ov-tcd-badge ${{tcdOk?'ok':'bad'}}">${{tcdOk?'✅ BATEU TCD':'⚠️ NÃO BATEU'}}</div></div>`;

  h+=`<div class="ov-cct-grid"><div><div class="ov-col-hdr corte" style="cursor:pointer" onclick="switchTab(4)" title="Clique para ver detalhes">✂️ CORTE <span style="font-size:10px;opacity:.6;margin-left:auto">ver detalhe →</span></div>${{(d.corte||[]).map(kpiCard).join('')}}</div><div><div class="ov-col-hdr carreg">🚜 CARREGAMENTO</div>${{(d.carreg||[]).map(kpiCard).join('')}}</div><div><div class="ov-col-hdr transp" style="cursor:pointer" onclick="switchTab(5)" title="Clique para ver detalhes">🚚 TRANSPORTE <span style="font-size:10px;opacity:.6;margin-left:auto">ver detalhe →</span></div>${{(d.transp||[]).map(kpiCard).join('')}}</div></div>`;

  document.getElementById('overviewContent').innerHTML=h;
}}


// ═══════ DETALHE CORTE (Tab 4) ═══════
function renderCorteDetail(){{
  if(!OV||!OV.corte_fr)return;
  const minDt=OV.dates&&OV.dates.length?OV.dates[0]:'';
  const maxDt=OV.dates&&OV.dates.length?OV.dates[OV.dates.length-1]:'';

  let h='<div class="p-filters" style="margin-bottom:16px"><div class="p-filter-label">Frente</div>';
  OV.frentes.forEach(f=>{{const act=f===ovFrente?'active':'';const lbl=f==='todas'?'Todas':'Frente '+f;h+=`<button class="p-filter-btn ${{act}}" onclick="setCorteFrente('${{f}}')">${{lbl}}</button>`}});
  h+=`<div class="p-filter-label" style="margin-left:16px">Período</div><input type="date" id="corteDtS" class="p-filter-btn" value="${{ovDtStart}}" min="${{minDt}}" max="${{maxDt}}" onchange="ovDtStart=this.value;corteRendered=false;renderCorteDetail();corteRendered=true"><span style="font-size:11px;color:var(--tx2);padding:0 4px">até</span><input type="date" id="corteDtE" class="p-filter-btn" value="${{ovDtEnd}}" min="${{minDt}}" max="${{maxDt}}" onchange="ovDtEnd=this.value;corteRendered=false;renderCorteDetail();corteRendered=true">`;
  h+='</div>';

  const cfList=ovFrente==='todas'?OV.corte_fr:OV.corte_fr.filter(c=>c.frente===ovFrente);
  h+='<div class="ov-cct-grid">';
  cfList.forEach(cf=>{{
    const mnOk=cf.manobraReal<=cf.manobraMeta||cf.manobraReal===0;
    h+=`<div class="ov-panel" style="margin-bottom:0"><div style="font-size:14px;font-weight:800;margin-bottom:12px;display:flex;align-items:center;gap:6px"><span class="frb frb-${{cf.frente}}">${{cf.frente}}</span> Frente ${{cf.frente}}</div>`;
    h+=`<div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:10px">`;
    h+=`<div class="ov-kpi ok"><div><div class="ov-kpi-nm">Vel. Corte</div><div class="ov-kpi-vl">${{fmt(cf.velCorte,2)}}</div><div class="ov-kpi-mt">km/h</div></div></div>`;
    h+=`<div class="ov-kpi ${{mnOk?'ok':'bad'}}"><div><div class="ov-kpi-nm">Manobra</div><div class="ov-kpi-vl">${{fmt(cf.manobraReal,1)}}%</div><div class="ov-kpi-mt">Meta: <b>${{fmt(cf.manobraMeta,1)}}%</b></div></div></div>`;
    h+=`<div class="ov-kpi ok"><div><div class="ov-kpi-nm">ATR</div><div class="ov-kpi-vl">${{fmt(cf.atr,1)}}</div><div class="ov-kpi-mt">kg/ton</div></div></div>`;
    h+=`<div class="ov-kpi ok"><div><div class="ov-kpi-nm">Chuva</div><div class="ov-kpi-vl">${{fmt(cf.chuvaHMaq,1)}}</div><div class="ov-kpi-mt">h/maq</div></div></div>`;
    h+=`<div class="ov-kpi ${{cf.hrsEfReal>=cf.hrsEfMeta||cf.hrsEfReal===0?'ok':'bad'}}"><div><div class="ov-kpi-nm">Hrs Efetivo</div><div class="ov-kpi-vl">${{fmt(cf.hrsEfReal,1)}}</div><div class="ov-kpi-mt">Meta: <b>${{fmt(cf.hrsEfMeta,1)}} h</b></div></div></div>`;
    h+=`<div class="ov-kpi ok"><div><div class="ov-kpi-nm">% Corte</div><div class="ov-kpi-vl">${{fmt(cf.pctCorte,1)}}%</div><div class="ov-kpi-mt">do tempo total</div></div></div>`;
    h+=`</div>`;
    h+=`<div style="font-size:10px;color:var(--tx2);display:grid;grid-template-columns:repeat(3,1fr);gap:4px;text-align:center">`;
    h+=`<div>Corte <b>${{fmt(cf.corteH,1)}}h</b></div><div>Manobra <b>${{fmt(cf.manobraH,1)}}h</b></div><div>Chuva <b>${{fmt(cf.chuvaH,1)}}h</b></div>`;
    h+=`<div>Limit.Ind <b>${{fmt(cf.limitIndH,1)}}h</b></div><div>Manut <b>${{fmt(cf.manutH,1)}}h</b></div><div>Pátio <b>${{fmt(cf.patioH,1)}}h</b></div>`;
    h+=`</div></div>`;
  }});
  h+='</div>';

  h+=`<div class="ov-panel"><div class="ov-panel-title"><div class="dot" style="background:var(--rz)"></div>Distribuição de Horas por Colhedora (Frota)</div><div class="chart-ww" style="height:350px"><canvas id="ovCorteChart2"></canvas></div></div>`;

  if(OV.raw_bordo_cd&&OV.raw_bordo_cd.length>0){{
    const rows=OV.raw_bordo_cd.filter(r=>(ovFrente==='todas'||r.fr===ovFrente) && (!ovDtStart||r.dt>=ovDtStart) && (!ovDtEnd||r.dt<=ovDtEnd));
    const agg={{}};
    rows.forEach(r=>{{
      const k=r.eq+'|'+r.fr;
      if(!agg[k])agg[k]={{eq:r.eq,fr:r.fr,co:0,mn:0,ds:0,ch:0,li:0,ma:0,pa:0,sa:0,tot:0,vpC:0,hoC:0}};
      const a=agg[k];a.co+=r.co;a.mn+=r.mn;a.ds+=r.ds;a.ch+=r.ch;a.li+=r.li;a.ma+=r.ma;a.pa+=r.pa;a.sa+=r.sa;a.tot+=r.tot;a.vpC+=r.vpC;a.hoC+=r.hoC;
    }});
    const list=Object.values(agg).filter(a=>a.tot>0).sort((a,b)=>b.tot-a.tot);
    h+=`<div class="ov-panel"><div class="ov-panel-title"><div class="dot" style="background:var(--bl)"></div>Ofensores por Colhedora (Frota) — Detalhe de Horas</div><div style="overflow-x:auto"><table class="ov-fleet"><thead><tr><th>Frota</th><th>Frente</th><th style="background:#1a6b30">Corte (h)</th><th style="background:#e17055">Manobra (h)</th><th>Desloc. (h)</th><th style="background:#74b9ff">Chuva (h)</th><th style="background:#fdcb6e">Limit.Ind (h)</th><th style="background:#d63031">Manut (h)</th><th style="background:#636e72">Pátio (h)</th><th>S/Apont (h)</th><th>Total (h)</th><th>Vel Corte</th><th>% Corte</th><th>% Manobra</th></tr></thead><tbody>`;
    list.forEach(a=>{{
      const vel=a.hoC>0?(a.vpC/a.hoC).toFixed(2):'—';
      const pCorte=a.tot>0?(a.co/a.tot*100).toFixed(1):'0';
      const pMan=a.tot>0?(a.mn/a.tot*100).toFixed(1):'0';
      h+=`<tr><td class="equip-cell">${{a.eq}}</td><td><span class="frb frb-${{a.fr}}">${{a.fr}}</span></td><td style="color:var(--rz);font-weight:700">${{fmt(a.co,1)}}</td><td style="color:var(--am);font-weight:700">${{fmt(a.mn,1)}}</td><td>${{fmt(a.ds,1)}}</td><td>${{fmt(a.ch,1)}}</td><td>${{fmt(a.li,1)}}</td><td style="color:var(--red)">${{fmt(a.ma,1)}}</td><td>${{fmt(a.pa,1)}}</td><td>${{fmt(a.sa,1)}}</td><td style="font-weight:700">${{fmt(a.tot,1)}}</td><td>${{vel}}</td><td style="color:var(--rz);font-weight:600">${{pCorte}}%</td><td style="color:${{parseFloat(pMan)>20?'var(--red)':'var(--tx)'}};font-weight:600">${{pMan}}%</td></tr>`;
    }});
    h+=`</tbody></table></div></div>`;
  }}

  if(OV.raw_op_cd&&OV.raw_op_cd.length>0){{
    const rowsOp=OV.raw_op_cd.filter(r=>(ovFrente==='todas'||r.fr===ovFrente) && (!ovDtStart||r.dt>=ovDtStart) && (!ovDtEnd||r.dt<=ovDtEnd));
    const agg={{}};
    rowsOp.forEach(r=>{{
      const k=r.op+'|'+r.fr;
      if(!agg[k])agg[k]={{op:r.op,fr:r.fr,frotas:{{}},co:0,mn:0,ds:0,ch:0,li:0,ma:0,pa:0,sa:0,ot:0,tot:0,vpC:0,hoC:0}};
      const a=agg[k];
      a.frotas[r.eq]=1;
      const d=r.desc;const h2=r.h;
      if(d==='CORTE DE CANA MECANIZADO'){{a.co+=h2;a.vpC+=r.vpC;a.hoC+=r.hoC}}
      else if(d==='MANOBRA')a.mn+=h2;
      else if(d==='DESLOCAMENTO')a.ds+=h2;
      else if(d.indexOf('CLIMATICA')>-1||d==='TEMPO INDETERMINADO')a.ch+=h2;
      else if(d==='LIMITAÇÃO INDÚSTRIA')a.li+=h2;
      else if(d.indexOf('MAN CORRETIVA')>-1||d.indexOf('AGUARDANDO MANUT')>-1||d==='TROCA DE FAQUINHA'||d.indexOf('LIMPEZA')>-1)a.ma+=h2;
      else if(d==='PATIO - RESERVA')a.pa+=h2;
      else if(d==='SEM APONTAMENTO')a.sa+=h2;
      else a.ot+=h2;
      a.tot+=h2;
    }});
    const list=Object.values(agg).filter(a=>a.tot>0).sort((a,b)=>b.tot-a.tot);
    h+=`<div class="ov-panel" id="secOpRanking"><div class="ov-panel-title"><div class="dot" style="background:var(--pr)"></div>🏆 Ranking por Operador — Detalhe de Horas</div><div style="overflow-x:auto"><table class="ov-fleet"><thead><tr><th>Operador</th><th>Frente</th><th>Frotas</th><th style="background:#1a6b30">Corte (h)</th><th style="background:#e17055">Manobra (h)</th><th>Desloc. (h)</th><th style="background:#74b9ff">Chuva (h)</th><th style="background:#fdcb6e">Limit.Ind (h)</th><th style="background:#d63031">Manut (h)</th><th style="background:#636e72">Pátio (h)</th><th>S/Apont (h)</th><th>Outros (h)</th><th>Total (h)</th><th>Vel Corte</th><th>% Corte</th><th>% Manobra</th></tr></thead><tbody>`;
    list.forEach(a=>{{
      const vel=a.hoC>0?(a.vpC/a.hoC).toFixed(2):'—';
      const pCorte=a.tot>0?(a.co/a.tot*100).toFixed(1):'0';
      const pMan=a.tot>0?(a.mn/a.tot*100).toFixed(1):'0';
      const frs=Object.keys(a.frotas).join(', ');
      h+=`<tr><td class="equip-cell" style="font-weight:700">${{a.op}}</td><td><span class="frb frb-${{a.fr}}">${{a.fr}}</span></td><td style="font-size:10px">${{frs}}</td><td style="color:var(--rz);font-weight:700">${{fmt(a.co,1)}}</td><td style="color:var(--am);font-weight:700">${{fmt(a.mn,1)}}</td><td>${{fmt(a.ds,1)}}</td><td>${{fmt(a.ch,1)}}</td><td>${{fmt(a.li,1)}}</td><td style="color:var(--red)">${{fmt(a.ma,1)}}</td><td>${{fmt(a.pa,1)}}</td><td>${{fmt(a.sa,1)}}</td><td>${{fmt(a.ot,1)}}</td><td style="font-weight:700">${{fmt(a.tot,1)}}</td><td>${{vel}}</td><td style="color:var(--rz);font-weight:600">${{pCorte}}%</td><td style="color:${{parseFloat(pMan)>20?'var(--red)':'var(--tx)'}};font-weight:600">${{pMan}}%</td></tr>`;
    }});
    h+=`</tbody></table></div></div>`;
  }}

  document.getElementById('corteContent').innerHTML=h;
  renderOvCorteChart2();
}}
function setCorteFrente(f){{ovFrente=f;corteRendered=false;renderCorteDetail();corteRendered=true}}

// ═══════ DETALHE TRANSPORTE (Tab 5) ═══════
function renderTransporteDetail(){{
  if(!OV||!OV.fleet)return;
  const minDt=OV.dates&&OV.dates.length?OV.dates[0]:'';
  const maxDt=OV.dates&&OV.dates.length?OV.dates[OV.dates.length-1]:'';

  const M=OV.metas;const fl=getOvFleet(ovFrente);
  const avg=(arr,k)=>arr.length?arr.reduce((s,r)=>s+r[k],0)/arr.length:0;
  const gauges=[
    {{label:'Vel. Vazio (km/h)',val:avg(fl,'velVazio'),meta:M.velVazio,hi:true}},
    {{label:'Vel. Carreg. (km/h)',val:avg(fl,'velCarreg'),meta:M.velCarreg,hi:true}},
    {{label:'Raio (km)',val:avg(fl,'raio'),meta:M.raio,hi:false,neutral:true}},
    {{label:'Densidade (t/v)',val:avg(fl,'densidade'),meta:M.densidade,hi:true}},
  ];

  let h='<div class="p-filters" style="margin-bottom:16px"><div class="p-filter-label">Frente</div>';
  OV.frentes.forEach(f=>{{const act=f===ovFrente?'active':'';const lbl=f==='todas'?'Todas':'Frente '+f;h+=`<button class="p-filter-btn ${{act}}" onclick="setTranspFrente('${{f}}')">${{lbl}}</button>`}});
  h+=`<div class="p-filter-label" style="margin-left:16px">Período</div><input type="date" id="transpDtS" class="p-filter-btn" value="${{ovDtStart}}" min="${{minDt}}" max="${{maxDt}}" onchange="ovDtStart=this.value;transpRendered=false;renderTransporteDetail();transpRendered=true"><span style="font-size:11px;color:var(--tx2);padding:0 4px">até</span><input type="date" id="transpDtE" class="p-filter-btn" value="${{ovDtEnd}}" min="${{minDt}}" max="${{maxDt}}" onchange="ovDtEnd=this.value;transpRendered=false;renderTransporteDetail();transpRendered=true">`;
  h+='</div>';

  h+='<div class="ov-gauge-row">';gauges.forEach(g=>{{const ok=g.neutral?true:(g.hi?g.val>=g.meta:g.val<=g.meta);h+=`<div class="ov-gauge ${{ok?'ok':'bad'}}"><div class="ov-gauge-lbl">${{g.label}}</div><div class="ov-gauge-val">${{fmt(g.val,1)}}</div><div class="ov-gauge-meta">Meta: ${{fmt(g.meta,1)}}</div></div>`}});h+='</div>';

  h+=`<div class="cg2"><div class="ov-panel"><div class="ov-panel-title"><div class="dot" style="background:var(--rz)"></div>Vel. Vazio — Ranking Frota</div><div class="chart-ww"><canvas id="tVelV"></canvas></div></div><div class="ov-panel"><div class="ov-panel-title"><div class="dot" style="background:var(--bl)"></div>Vel. Carregado — Ranking Frota</div><div class="chart-ww"><canvas id="tVelC"></canvas></div></div></div>`;

  h+=`<div class="ov-panel"><div class="ov-panel-title"><div class="dot" style="background:var(--bl)"></div>Visão Detalhada da Frota</div><div style="overflow-x:auto"><table class="ov-fleet" id="tFleet"><thead><tr><th rowspan="2">Equip.</th><th rowspan="2">Fr.</th><th colspan="2" style="background:#1a6b30">Velocidade</th><th colspan="2" style="background:#7b5e00">Ciclo (min)</th><th rowspan="2" class="sort" data-col="raio">Raio</th><th rowspan="2" class="sort" data-col="densidade">Dens.</th><th colspan="4" style="background:#8b1a1a">Ofensores (h)</th><th rowspan="2" class="sort" data-col="pctOfensor">%Of.</th></tr><tr><th class="sort" data-col="velVazio" style="background:#228B3E">Vazio</th><th class="sort" data-col="velCarreg" style="background:#228B3E">Carr.</th><th class="sort" data-col="t1" style="background:#9a7d00">T1</th><th class="sort" data-col="t2" style="background:#9a7d00">T2</th><th class="sort" data-col="manut" style="background:#a52020">Man.</th><th class="sort" data-col="agTransb" style="background:#a52020">Ag.T</th><th class="sort" data-col="semApont" style="background:#a52020">S/A</th><th class="sort" data-col="patio" style="background:#a52020">Pat.</th></tr></thead><tbody></tbody></table></div></div>`;

  h+=`<div class="ov-panel"><div class="ov-panel-title"><div class="dot" style="background:var(--am)"></div>Ofensores por Equipamento — Manutenção, Ag.Transbordo, Sem Apont., Pátio (horas)</div><div class="chart-ww"><canvas id="tOfChart"></canvas></div></div>`;

  document.getElementById('transporteContent').innerHTML=h;

  document.querySelector('#tFleet thead').addEventListener('click',e=>{{const th=e.target.closest('th.sort');if(!th)return;const c=th.dataset.col;if(ovSortCol===c)ovSortDir=ovSortDir==='asc'?'desc':'asc';else{{ovSortCol=c;ovSortDir='desc'}};renderTranspFleetTable()}});

  renderTranspFleetTable();
  renderTranspCharts();
}}
function setTranspFrente(f){{ovFrente=f;transpRendered=false;renderTransporteDetail();transpRendered=true}}

function renderTranspFleetTable(){{
  const fl=getOvFleet(ovFrente);const M=OV.metas;
  fl.sort((a,b)=>ovSortDir==='asc'?a[ovSortCol]-b[ovSortCol]:b[ovSortCol]-a[ovSortCol]);
  const tb=document.querySelector('#tFleet tbody');
  let h=`<tr class="meta-row"><td colspan="2" style="text-align:right;font-size:9px">META</td><td>${{fmt(M.velVazio,0)}}</td><td>${{fmt(M.velCarreg,0)}}</td><td>≤${{M.t1}}</td><td>≤${{fmt(M.t2,0)}}</td><td style="color:var(--tx2)">${{fmt(M.raio,0)}}</td><td>≥${{fmt(M.densidade,0)}}</td><td colspan="4">—</td><td>≤${{fmt(M.pctOfensor,0)}}%</td></tr>`;
  h+=fl.map(r=>{{return`<tr><td class="equip-cell">${{r.equip}}</td><td><span class="frb frb-${{r.frente}}">${{r.frente}}</span></td><td class="${{r.velVazio>=M.velVazio?'vp':'vn'}}">${{fmt(r.velVazio,1)}}</td><td class="${{r.velCarreg>=M.velCarreg?'vp':'vn'}}">${{fmt(r.velCarreg,1)}}</td><td class="${{r.t1<=M.t1?'vp':'vn'}}">${{fmt(r.t1,0)}}</td><td class="${{r.t2<=M.t2?'vp':'vn'}}">${{fmt(r.t2,0)}}</td><td>${{fmt(r.raio,1)}}</td><td class="${{r.densidade>=M.densidade?'vp':'vn'}}">${{fmt(r.densidade,1)}}</td><td>${{fmt(r.manut,1)}}</td><td>${{fmt(r.agTransb,1)}}</td><td>${{fmt(r.semApont,1)}}</td><td>${{fmt(r.patio,1)}}</td><td><div class="bar-c"><span style="font-weight:700;min-width:34px;font-size:10px;color:${{r.pctOfensor>M.pctOfensor?'var(--red)':'var(--tx)'}}">${{fmt(r.pctOfensor,0)}}%</span><div class="bar-bg"><div class="bar-f ${{r.pctOfensor>M.pctOfensor?'high':'low'}}" style="width:${{Math.min(r.pctOfensor,100)}}%"></div></div></div></td></tr>`}}).join('');
  tb.innerHTML=h;
  document.querySelectorAll('#tFleet th.sort').forEach(th=>{{th.classList.remove('asc','desc');if(th.dataset.col===ovSortCol)th.classList.add(ovSortDir)}});
}}

function renderTranspCharts(){{
  const fl=getOvFleet(ovFrente);const M=OV.metas;const JK="'Plus Jakarta Sans'";
  const dlCfg={{display:(ctx)=>ctx.parsed&&ctx.datasetIndex===0,anchor:'end',align:'top',color:'#1a2e22',font:{{family:'JetBrains Mono',size:9,weight:'bold'}},formatter:v=>v?fmt(v,1):''}};
  const sv=fl.slice().sort((a,b)=>b.velVazio-a.velVazio);
  mkOvChart('tVelV',{{type:'bar',data:{{labels:sv.map(r=>r.equip),datasets:[{{data:sv.map(r=>r.velVazio),backgroundColor:sv.map(r=>r.velVazio>=M.velVazio?'#00843D':'#d63031'),borderRadius:3}},{{type:'line',data:sv.map(()=>M.velVazio),borderColor:'#1a1a2e',borderWidth:1.5,borderDash:[5,3],pointRadius:0}}]}},options:{{responsive:true,maintainAspectRatio:false,layout:{{padding:{{top:22}}}},plugins:{{legend:{{display:false}},datalabels:dlCfg}},scales:{{y:{{beginAtZero:true,grid:{{color:'#eee'}}}},x:{{ticks:{{font:{{family:JK,size:9,weight:'600'}}}},grid:{{display:false}}}}}}}}}});
  const sc=fl.slice().sort((a,b)=>b.velCarreg-a.velCarreg);
  mkOvChart('tVelC',{{type:'bar',data:{{labels:sc.map(r=>r.equip),datasets:[{{data:sc.map(r=>r.velCarreg),backgroundColor:sc.map(r=>r.velCarreg>=M.velCarreg?'#0984e3':'#d63031'),borderRadius:3}},{{type:'line',data:sc.map(()=>M.velCarreg),borderColor:'#1a1a2e',borderWidth:1.5,borderDash:[5,3],pointRadius:0}}]}},options:{{responsive:true,maintainAspectRatio:false,layout:{{padding:{{top:22}}}},plugins:{{legend:{{display:false}},datalabels:dlCfg}},scales:{{y:{{beginAtZero:true,grid:{{color:'#eee'}}}},x:{{ticks:{{font:{{family:JK,size:9,weight:'600'}}}},grid:{{display:false}}}}}}}}}});
  const so=fl.slice().sort((a,b)=>(b.manut+b.agTransb+b.patio)-(a.manut+a.agTransb+a.patio));
  mkOvChart('tOfChart',{{type:'bar',data:{{labels:so.map(r=>r.equip),datasets:[{{label:'Manutenção',data:so.map(r=>r.manut),backgroundColor:'#d63031',borderRadius:3}},{{label:'Ag. Transbordo',data:so.map(r=>r.agTransb),backgroundColor:'#e17055',borderRadius:3}},{{label:'Sem Apontamento',data:so.map(r=>r.semApont),backgroundColor:'#fdcb6e',borderRadius:3}},{{label:'Pátio',data:so.map(r=>r.patio),backgroundColor:'#636e72',borderRadius:3}}]}},options:{{responsive:true,maintainAspectRatio:false,scales:{{x:{{stacked:true,ticks:{{font:{{family:JK,size:9,weight:'600'}}}},grid:{{display:false}}}},y:{{stacked:true,title:{{display:true,text:'Horas',font:{{family:JK,size:10}}}},grid:{{color:'#eee'}}}}}},plugins:{{legend:{{position:'top',labels:{{font:{{family:JK,size:10}},usePointStyle:true,pointStyleWidth:9,padding:12}}}},datalabels:{{display:(ctx)=>ctx.parsed&&ctx.parsed.y>=1.5,anchor:'center',align:'center',color:'#fff',font:{{family:'JetBrains Mono',size:9,weight:'bold'}},formatter:(v)=>v>=1.5?fmt(v,1):''}},tooltip:{{callbacks:{{label:c=>' '+c.dataset.label+': '+fmt(c.parsed.y,1)+'h'}}}}}}}}}});
}}

function renderOvCorteChart2(){{
  if(!OV.raw_bordo_cd||!OV.raw_bordo_cd.length)return;
  const rows=OV.raw_bordo_cd.filter(r=>(ovFrente==='todas'||r.fr===ovFrente) && (!ovDtStart||r.dt>=ovDtStart) && (!ovDtEnd||r.dt<=ovDtEnd));
  const agg={{}};
  rows.forEach(r=>{{
    if(!agg[r.eq])agg[r.eq]={{eq:r.eq,fr:r.fr,co:0,mn:0,ds:0,ch:0,li:0,ma:0,pa:0,sa:0}};
    const a=agg[r.eq];a.co+=r.co;a.mn+=r.mn;a.ds+=r.ds;a.ch+=r.ch;a.li+=r.li;a.ma+=r.ma;a.pa+=r.pa;a.sa+=r.sa;
  }});
  const list=Object.values(agg).sort((a,b)=>(b.co+b.mn+b.ds+b.ch+b.li+b.ma+b.pa+b.sa)-(a.co+a.mn+a.ds+a.ch+a.li+a.ma+a.pa+a.sa));
  if(!list.length)return;
  const JK="'Plus Jakarta Sans'";
  const labels=list.map(a=>a.eq+' ('+a.fr+')');
  mkOvChart('ovCorteChart2',{{type:'bar',data:{{labels,datasets:[
    {{label:'Corte',data:list.map(a=>a.co),backgroundColor:'#00843D',borderRadius:3}},
    {{label:'Manobra',data:list.map(a=>a.mn),backgroundColor:'#e17055',borderRadius:3}},
    {{label:'Deslocamento',data:list.map(a=>a.ds),backgroundColor:'#0984e3',borderRadius:3}},
    {{label:'Chuva',data:list.map(a=>a.ch),backgroundColor:'#74b9ff',borderRadius:3}},
    {{label:'Limit.Ind',data:list.map(a=>a.li),backgroundColor:'#fdcb6e',borderRadius:3}},
    {{label:'Manut',data:list.map(a=>a.ma),backgroundColor:'#d63031',borderRadius:3}},
    {{label:'Pátio',data:list.map(a=>a.pa),backgroundColor:'#636e72',borderRadius:3}},
    {{label:'S/Apont',data:list.map(a=>a.sa),backgroundColor:'#a29bfe',borderRadius:3}},
  ]}},options:{{responsive:true,maintainAspectRatio:false,layout:{{padding:{{top:20}}}},scales:{{x:{{stacked:true,ticks:{{font:{{family:JK,size:9,weight:'600'}}}},grid:{{display:false}}}},y:{{stacked:true,title:{{display:true,text:'Horas',font:{{family:JK,size:10}}}},grid:{{color:'#eee'}}}}}},plugins:{{legend:{{position:'top',labels:{{font:{{family:JK,size:10}},usePointStyle:true,pointStyleWidth:9,padding:12}}}},datalabels:{{display:(ctx)=>ctx.parsed&&ctx.parsed.y>=3,anchor:'center',align:'center',color:'#fff',font:{{family:'JetBrains Mono',size:8,weight:'bold'}},formatter:v=>v>=3?fmt(v,0)+'h':''}},tooltip:{{callbacks:{{label:c=>' '+c.dataset.label+': '+fmt(c.parsed.y,1)+'h'}}}}}}}}}});
}}

function renderOvCorteChart(){{
  if(!OV.corte_fr||!OV.corte_fr.length)return;
  const cf=OV.corte_fr;const JK="'Plus Jakarta Sans'";
  const labels=cf.map(c=>'Fr '+c.frente);
  mkOvChart('ovCorteChart',{{type:'bar',data:{{labels,datasets:[
    {{label:'Corte',data:cf.map(c=>c.corteH),backgroundColor:'#00843D',borderRadius:3}},
    {{label:'Manobra',data:cf.map(c=>c.manobraH),backgroundColor:'#e17055',borderRadius:3}},
    {{label:'Chuva/Clima',data:cf.map(c=>c.chuvaH),backgroundColor:'#74b9ff',borderRadius:3}},
    {{label:'Limit. Indústria',data:cf.map(c=>c.limitIndH),backgroundColor:'#fdcb6e',borderRadius:3}},
    {{label:'Manutenção',data:cf.map(c=>c.manutH),backgroundColor:'#d63031',borderRadius:3}},
    {{label:'Pátio',data:cf.map(c=>c.patioH),backgroundColor:'#636e72',borderRadius:3}},
    {{label:'Sem Apont.',data:cf.map(c=>c.semApontH),backgroundColor:'#a29bfe',borderRadius:3}},
  ]}},options:{{responsive:true,maintainAspectRatio:false,layout:{{padding:{{top:20}}}},scales:{{x:{{stacked:true,ticks:{{font:{{family:JK,size:11,weight:'600'}}}},grid:{{display:false}}}},y:{{stacked:true,title:{{display:true,text:'Horas',font:{{family:JK,size:10}}}},grid:{{color:'#eee'}}}}}},plugins:{{legend:{{position:'top',labels:{{font:{{family:JK,size:10}},usePointStyle:true,pointStyleWidth:9,padding:12}}}},datalabels:{{display:(ctx)=>ctx.parsed&&ctx.parsed.y>=5,anchor:'center',align:'center',color:'#fff',font:{{family:'JetBrains Mono',size:9,weight:'bold'}},formatter:v=>v>=5?fmt(v,0)+'h':''}},tooltip:{{callbacks:{{label:c=>' '+c.dataset.label+': '+fmt(c.parsed.y,1)+'h'}}}}}}}}}});
}}

function setOvFrente(f){{ovFrente=f;ovRendered=false;renderOverviewFull();ovRendered=true}}

function renderOvFleetTable(){{
  const fl=getOvFleet(ovFrente);const M=OV.metas;
  fl.sort((a,b)=>ovSortDir==='asc'?a[ovSortCol]-b[ovSortCol]:b[ovSortCol]-a[ovSortCol]);
  const tb=document.querySelector('#ovFleet tbody');
  let h=`<tr class="meta-row"><td colspan="2" style="text-align:right;font-size:9px">META</td><td>${{fmt(M.velVazio,0)}}</td><td>${{fmt(M.velCarreg,0)}}</td><td>≤${{M.t1}}</td><td>≤${{fmt(M.t2,0)}}</td><td style="color:var(--tx2)">${{fmt(M.raio,0)}}</td><td>≥${{fmt(M.densidade,0)}}</td><td colspan="4">—</td><td>≤${{fmt(M.pctOfensor,0)}}%</td></tr>`;
  h+=fl.map(r=>{{return`<tr><td class="equip-cell">${{r.equip}}</td><td><span class="frb frb-${{r.frente}}">${{r.frente}}</span></td><td class="${{r.velVazio>=M.velVazio?'vp':'vn'}}">${{fmt(r.velVazio,1)}}</td><td class="${{r.velCarreg>=M.velCarreg?'vp':'vn'}}">${{fmt(r.velCarreg,1)}}</td><td class="${{r.t1<=M.t1?'vp':'vn'}}">${{fmt(r.t1,0)}}</td><td class="${{r.t2<=M.t2?'vp':'vn'}}">${{fmt(r.t2,0)}}</td><td>${{fmt(r.raio,1)}}</td><td class="${{r.densidade>=M.densidade?'vp':'vn'}}">${{fmt(r.densidade,1)}}</td><td>${{fmt(r.manut,1)}}</td><td>${{fmt(r.agTransb,1)}}</td><td>${{fmt(r.semApont,1)}}</td><td>${{fmt(r.patio,1)}}</td><td><div class="bar-c"><span style="font-weight:700;min-width:34px;font-size:10px;color:${{r.pctOfensor>M.pctOfensor?'var(--red)':'var(--tx)'}}">${{fmt(r.pctOfensor,0)}}%</span><div class="bar-bg"><div class="bar-f ${{r.pctOfensor>M.pctOfensor?'high':'low'}}" style="width:${{Math.min(r.pctOfensor,100)}}%"></div></div></div></td></tr>`}}).join('');
  tb.innerHTML=h;
  document.querySelectorAll('#ovFleet th.sort').forEach(th=>{{th.classList.remove('asc','desc');if(th.dataset.col===ovSortCol)th.classList.add(ovSortDir)}});
}}

function renderOvCharts(){{
  const fl=getOvFleet(ovFrente);const M=OV.metas;const JK="'Plus Jakarta Sans'";
  const sv=fl.slice().sort((a,b)=>b.velVazio-a.velVazio);
  mkOvChart('ovVelV',{{type:'bar',data:{{labels:sv.map(r=>r.equip),datasets:[{{data:sv.map(r=>r.velVazio),backgroundColor:sv.map(r=>r.velVazio>=M.velVazio?'#00843D':'#d63031'),borderRadius:3}},{{type:'line',data:sv.map(()=>M.velVazio),borderColor:'#1a1a2e',borderWidth:1.5,borderDash:[5,3],pointRadius:0,label:'Meta'}}]}},options:{{responsive:true,maintainAspectRatio:false,layout:{{padding:{{top:22}}}},plugins:{{legend:{{display:false}},datalabels:{{display:(ctx)=>ctx.datasetIndex===0,anchor:'end',align:'top',color:'#1a2e22',font:{{family:'JetBrains Mono',size:9,weight:'bold'}},formatter:v=>v?fmt(v,1):''}}}},scales:{{y:{{beginAtZero:true,grid:{{color:'#eee'}}}},x:{{ticks:{{font:{{family:JK,size:9,weight:'600'}}}},grid:{{display:false}}}}}}}}}});
  const sc=fl.slice().sort((a,b)=>b.velCarreg-a.velCarreg);
  mkOvChart('ovVelC',{{type:'bar',data:{{labels:sc.map(r=>r.equip),datasets:[{{data:sc.map(r=>r.velCarreg),backgroundColor:sc.map(r=>r.velCarreg>=M.velCarreg?'#0984e3':'#d63031'),borderRadius:3}},{{type:'line',data:sc.map(()=>M.velCarreg),borderColor:'#1a1a2e',borderWidth:1.5,borderDash:[5,3],pointRadius:0,label:'Meta'}}]}},options:{{responsive:true,maintainAspectRatio:false,layout:{{padding:{{top:22}}}},plugins:{{legend:{{display:false}},datalabels:{{display:(ctx)=>ctx.datasetIndex===0,anchor:'end',align:'top',color:'#1a2e22',font:{{family:'JetBrains Mono',size:9,weight:'bold'}},formatter:v=>v?fmt(v,1):''}}}},scales:{{y:{{beginAtZero:true,grid:{{color:'#eee'}}}},x:{{ticks:{{font:{{family:JK,size:9,weight:'600'}}}},grid:{{display:false}}}}}}}}}});
  const so=fl.slice().sort((a,b)=>(b.manut+b.agTransb+b.semApont+b.patio)-(a.manut+a.agTransb+a.semApont+a.patio));
  mkOvChart('ovOfChart',{{type:'bar',data:{{labels:so.map(r=>r.equip),datasets:[{{label:'Manutenção',data:so.map(r=>r.manut),backgroundColor:'#d63031',borderRadius:3}},{{label:'Ag. Transbordo',data:so.map(r=>r.agTransb),backgroundColor:'#e17055',borderRadius:3}},{{label:'Sem Apontamento',data:so.map(r=>r.semApont),backgroundColor:'#fdcb6e',borderRadius:3}},{{label:'Pátio',data:so.map(r=>r.patio),backgroundColor:'#636e72',borderRadius:3}}]}},options:{{responsive:true,maintainAspectRatio:false,scales:{{x:{{stacked:true,ticks:{{font:{{family:JK,size:9,weight:'600'}}}},grid:{{display:false}}}},y:{{stacked:true,title:{{display:true,text:'Horas',font:{{family:JK,size:10}}}},grid:{{color:'#eee'}}}}}},plugins:{{legend:{{position:'top',labels:{{font:{{family:JK,size:10}},usePointStyle:true,pointStyleWidth:9,padding:12}}}},datalabels:{{display:(ctx)=>ctx.parsed&&ctx.parsed.y>=1.5,anchor:'center',align:'center',color:'#fff',font:{{family:'JetBrains Mono',size:9,weight:'bold'}},formatter:(v)=>v>=1.5?fmt(v,1):''}},tooltip:{{callbacks:{{label:c=>' '+c.dataset.label+': '+fmt(c.parsed.y,1)+'h'}}}}}}}}}});
}}



renderProd();
</script></body></html>'''
    return html


def _build_overview_html_block():
    return ""

# ============================================================
# MAIN
# ============================================================
def monitorar(no_auth=False):
    pasta = Path(__file__).parent.resolve()
    os.chdir(pasta)
    print("=" * 65)
    print("  DASHBOARD SAFRA 25/26 — PAINEL INTEGRADO")
    print("=" * 65)
    xlsx_prod = pasta / ARQUIVO_PRODUCAO
    xlsx_metricas = pasta / ARQUIVO_METRICAS
    if xlsx_prod.exists(): print(f"  [OK] Produção: {xlsx_prod.name}")
    else: print(f"  [AVISO] {ARQUIVO_PRODUCAO} não encontrado.")
    if xlsx_metricas.exists(): print(f"  [OK] Métricas: {xlsx_metricas.name}")
    else: print(f"  [AVISO] {ARQUIVO_METRICAS} não encontrado.")
    print("-" * 65)
    try:
        jp, fp = ("{}", [])
        jper = "null"
        jov = "null"
        if xlsx_prod.exists():
            jp, fp = extract_producao(str(xlsx_prod))
            jper = extract_perdas(str(xlsx_prod))
        if xlsx_metricas.exists():
            jov = extract_overview(str(xlsx_metricas))
        html = build_html(jp, fp, jper, jov)
        if not no_auth:
            html = encrypt_dashboard(html)
        with open(ARQUIVO_HTML, 'w', encoding='utf-8') as f:
            f.write(html)
        print(f"  ✅ {ARQUIVO_HTML} gerado com sucesso!")
        try:
            os.system(f'git add "{ARQUIVO_HTML}"')
            os.system(f'git commit -m "Auto update {time.strftime("%d/%m/%Y %H:%M")}"')
            os.system('git push')
        except: pass
        if ABRIR_BROWSER:
            try: os.startfile(os.path.abspath(ARQUIVO_HTML))
            except: pass
    except Exception as e:
        print(f"  [ERRO] {e}")
        traceback.print_exc()
    # Monitor loop
    mod_p = os.path.getmtime(xlsx_prod) if xlsx_prod.exists() else 0
    mod_m = os.path.getmtime(xlsx_metricas) if xlsx_metricas.exists() else 0
    print("-" * 65)
    try:
        while True:
            time.sleep(INTERVALO_SEG)
            np2 = os.path.getmtime(xlsx_prod) if xlsx_prod.exists() else 0
            nm2 = os.path.getmtime(xlsx_metricas) if xlsx_metricas.exists() else 0
            if np2 != mod_p or nm2 != mod_m:
                mod_p, mod_m = np2, nm2
                print(f"\n  [{time.strftime('%H:%M:%S')}] Atualização detectada!")
                time.sleep(3)
                try:
                    j1, f1 = ("{}", [])
                    j2, j3 = "null", "null"
                    if xlsx_prod.exists(): j1, f1 = extract_producao(str(xlsx_prod)); j2 = extract_perdas(str(xlsx_prod))
                    if xlsx_metricas.exists(): j3 = extract_overview(str(xlsx_metricas))
                    html = build_html(j1, f1, j2, j3)
                    if not no_auth:
                        html = encrypt_dashboard(html)
                    with open(ARQUIVO_HTML, 'w', encoding='utf-8') as f: f.write(html)
                    try: os.system(f'git add "{ARQUIVO_HTML}"'); os.system(f'git commit -m "Auto {time.strftime("%d/%m/%Y %H:%M")}"'); os.system('git push')
                    except: pass
                    print(f"  ✅ Atualizado!")
                except Exception as e: print(f"  [ERRO] {e}")
                print("-" * 65)
            else:
                print(f"\r  Monitorando... {time.strftime('%H:%M:%S')}", end="", flush=True)
    except KeyboardInterrupt:
        print("\n  Encerrado.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Dashboard CCT Ipaussu — Painel Integrado')
    parser.add_argument('--add', nargs=2, metavar=('USER','SENHA'), help='parser.add_argument('--add', nargs=2, metavar=('USER','SENHA'), help='Adicionar usuário')
    parser.add_argument('--remove', metavar='USER', help='Remover usuário')
    parser.add_argument('--users', action='store_true', help='Listar usuários')
    parser.add_argument('--no-auth', action='store_true', help='Gerar sem autenticação')
    args = parser.parse_args()

    if args.add:
        add_usuario(args.add[0].lower(), args.add[1])
    elif args.remove:
        del_usuario(args.remove.lower())
    elif args.users:
        list_usuarios()
    else:
        monitorar(no_auth=args.no_auth)