"""
╔══════════════════════════════════════════════════════════════╗
║  ROB CORRETIVO — Dashboard Controle de Estoque Corretivos    ║
║  Lê a planilha, gera HTML, e serve API para lançar NF        ║
╚══════════════════════════════════════════════════════════════╝

MODOS:
  python rob_corretivo.py              → Gera HTML estático
  python rob_corretivo.py --server     → Inicia servidor com lançamento de NF
  python rob_corretivo.py --server -p 8080  → Porta customizada
"""

import json, os, sys, time, re
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("Instale: pip install openpyxl"); sys.exit(1)

ARQUIVO_EXCEL = "Controle_corretivo_SF26'27.xlsx"
ARQUIVO_HTML  = "dash_corretivo.html"

def sf(v, d=2):
    if v is None or v == '-' or v == '': return 0
    try: return round(float(v), d)
    except: return 0

def ler_dados():
    wb = load_workbook(ARQUIVO_EXCEL, data_only=True)

    # ── 1) Controle de Estoque ──
    ws = wb['Controle de Estoque']
    estoque = []
    for r in range(4, ws.max_row+1):
        cod = ws.cell(r,1).value
        nome = ws.cell(r,2).value
        if not cod: continue
        estoque.append({
            'cod': str(cod),
            'nome': str(nome or ''),
            'zona': str(ws.cell(r,3).value or ''),
            'tipo': str(ws.cell(r,4).value or ''),
            'recCalcT': sf(ws.cell(r,5).value),
            'recCalcR': sf(ws.cell(r,6).value),
            'recGessoR': sf(ws.cell(r,7).value),
            'recGessoT': sf(ws.cell(r,8).value),
            'nfGesso': sf(ws.cell(r,9).value),
            'nfCalc': sf(ws.cell(r,10).value),
            'apCalc': sf(ws.cell(r,11).value),
            'apGesso': sf(ws.cell(r,12).value),
            'stCalc': str(ws.cell(r,13).value or ''),
            'stGesso': str(ws.cell(r,14).value or ''),
            'saldoCalc': sf(ws.cell(r,15).value),
            'saldoGesso': sf(ws.cell(r,16).value),
        })
    print(f"  [OK] Controle: {len(estoque)} registros")

    # ── 2) Base Notas ──
    ws2 = wb['Base Notas ']
    notas = []
    for r in range(6, ws2.max_row+1):
        nf = ws2.cell(r,2).value
        if not nf: continue
        dt = ws2.cell(r,3).value
        notas.append({
            'chave': str(ws2.cell(r,1).value or ''),
            'nf': str(int(float(nf))) if nf else '',
            'data': dt.strftime('%Y-%m-%d') if hasattr(dt,'strftime') else str(dt or '')[:10],
            'fazCod': str(int(float(ws2.cell(r,4).value))) if ws2.cell(r,4).value and str(ws2.cell(r,4).value).replace('.','').replace('-','').isdigit() else str(ws2.cell(r,4).value or ''),
            'fazNome': str(ws2.cell(r,5).value or ''),
            'zona': str(int(float(ws2.cell(r,6).value))) if ws2.cell(r,6).value and str(ws2.cell(r,6).value).replace('.','').replace('-','').isdigit() else str(ws2.cell(r,6).value or ''),
            'insumo': str(ws2.cell(r,7).value or ''),
            'peso': sf(ws2.cell(r,8).value),
            'status': ws2.cell(r,9).value.strftime('%Y-%m-%d') if hasattr(ws2.cell(r,9).value,'strftime') else str(ws2.cell(r,9).value or ''),
            'mes': str(ws2.cell(r,10).value or ''),
        })
    print(f"  [OK] Notas: {len(notas)} NFs")

    # ── 3) Tratos ──
    ws3 = wb['Tratos']
    tratos = []
    for r in range(2, ws3.max_row+1):
        num = ws3.cell(r,3).value
        if not num: continue
        tratos.append({
            'cod': str(int(float(num))),
            'fazenda': str(ws3.cell(r,4).value or ''),
            'zona': str(ws3.cell(r,5).value or ''),
            'talhao': str(ws3.cell(r,6).value or ''),
            'area': sf(ws3.cell(r,7).value),
            'varied': str(ws3.cell(r,11).value or ''),
            'estagio': str(ws3.cell(r,12).value or ''),
            'progCalc': str(ws3.cell(r,18).value or ''),
            'areaCalc': sf(ws3.cell(r,19).value),
            'dose': sf(ws3.cell(r,23).value),
            'volume': sf(ws3.cell(r,24).value),
        })
    print(f"  [OK] Tratos: {len(tratos)} talhões")

    # ── 4) Reforma ──
    ws4 = wb['Reforma']
    reforma = []
    for r in range(2, ws4.max_row+1):
        zona = ws4.cell(r,5).value
        if not zona: continue
        reforma.append({
            'zona': str(zona),
            'talhao': str(ws4.cell(r,6).value or ''),
            'tipo': str(ws4.cell(r,9).value or ''),
            'txCalc': sf(ws4.cell(r,10).value),
            'txGesso': sf(ws4.cell(r,11).value),
            'txFosf': sf(ws4.cell(r,12).value),
            'area': sf(ws4.cell(r,13).value),
            'qtdCalc': sf(ws4.cell(r,14).value),
            'qtdGesso': sf(ws4.cell(r,15).value),
            'qtdFosf': sf(ws4.cell(r,16).value),
        })
    print(f"  [OK] Reforma: {len(reforma)} talhões")

    # ── 5) Fazendas únicas para dropdown ──
    fazendas = []
    seen = set()
    for e in estoque:
        if e['tipo'] == 'Estoque': continue
        k = e['cod']
        if k not in seen:
            seen.add(k)
            fazendas.append({'cod':e['cod'],'nome':e['nome']})
    fazendas.sort(key=lambda x: x['nome'])

    return estoque, notas, tratos, reforma, fazendas

def inj(obj):
    return json.dumps(obj, separators=(",",":"), ensure_ascii=False).replace("</","<\\/")

def gerar_html(estoque, notas, tratos, reforma, fazendas):
    gerado = datetime.now().strftime("%d/%m/%Y %H:%M")
    L = []
    w = L.append

    w("<!DOCTYPE html><html lang='pt-BR'><head><meta charset='UTF-8'>")
    w("<meta name='viewport' content='width=device-width,initial-scale=1'>")
    w("<title>Controle Corretivos SF 26/27</title>")
    w('<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>')
    w('<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.2.0"></script>')
    w("<style>")
    w("""
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700;800&display=swap');
*{box-sizing:border-box;margin:0;padding:0}
:root{--bg:#f3f4f6;--card:#fff;--rz:#00843D;--rz2:#004D23;--bl:#1565C0;--am:#F59E0B;--red:#DC2626;--tx:#1a1a2e;--tx2:#6b7280;--brd:#e5e7eb;--ok:#10B981;--bad:#EF4444;--warn:#F59E0B}
body{font-family:'DM Sans',sans-serif;background:var(--bg);color:var(--tx);font-size:13px}
.topbar{background:linear-gradient(135deg,var(--rz2),var(--rz));color:#fff;padding:16px 24px;display:flex;align-items:center;justify-content:space-between}
.topbar h1{font-size:18px;font-weight:800;letter-spacing:-.02em}
.topbar .sub{font-size:11px;opacity:.7}
.tabs{display:flex;background:#fff;border-bottom:2px solid var(--brd);padding:0 16px;gap:0;overflow-x:auto}
.tab{padding:14px 22px;font-weight:700;font-size:13px;color:var(--tx2);border-bottom:3px solid transparent;cursor:pointer;white-space:nowrap;transition:all .2s}
.tab.on{color:var(--rz);border-bottom-color:var(--rz)}
.tab:hover{color:var(--rz)}
.pg{display:none;padding:16px 20px;max-width:1440px;margin:0 auto}.pg.on{display:block}
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px;margin-bottom:16px}
.kpi{background:var(--card);border-radius:12px;padding:16px;border-left:4px solid var(--rz);box-shadow:0 1px 3px rgba(0,0,0,.06)}
.kpi.bl{border-left-color:var(--bl)}.kpi.am{border-left-color:var(--am)}.kpi.red{border-left-color:var(--red)}.kpi.ok{border-left-color:var(--ok)}
.kpi-lb{font-size:10px;color:var(--tx2);text-transform:uppercase;font-weight:700;letter-spacing:.04em;margin-bottom:4px}
.kpi-vl{font-size:24px;font-weight:800}
.kpi-sub{font-size:10px;color:var(--tx2);margin-top:2px}
.card{background:var(--card);border-radius:12px;padding:16px;margin-bottom:14px;box-shadow:0 1px 3px rgba(0,0,0,.06)}
.card-title{font-size:14px;font-weight:800;margin-bottom:12px;display:flex;align-items:center;gap:8px}
.grid-2{display:grid;grid-template-columns:1fr 1fr;gap:14px}
table{width:100%;border-collapse:collapse;font-size:12px}
th{background:var(--rz);color:#fff;padding:10px 8px;text-align:left;font-weight:700;position:sticky;top:0;z-index:10}
td{padding:8px;border-bottom:1px solid var(--brd)}
tr:hover td{background:#f9fafb}
.num{text-align:right;font-variant-numeric:tabular-nums}
.st-ok{background:#d1fae5;color:#065f46;padding:2px 8px;border-radius:10px;font-size:10px;font-weight:700}
.st-bad{background:#fee2e2;color:#991b1b;padding:2px 8px;border-radius:10px;font-size:10px;font-weight:700}
.st-stock{background:#dbeafe;color:#1e40af;padding:2px 8px;border-radius:10px;font-size:10px;font-weight:700}
.st-na{background:#f3f4f6;color:#9ca3af;padding:2px 8px;border-radius:10px;font-size:10px;font-weight:700}
.filter-bar{display:flex;flex-wrap:wrap;gap:8px;margin-bottom:14px;align-items:center}
.filter-bar input,.filter-bar select{border:2px solid var(--brd);border-radius:8px;padding:8px 12px;font-size:12px;font-family:inherit}
.filter-bar input:focus,.filter-bar select:focus{border-color:var(--rz);outline:none}
.filter-bar label{font-size:11px;font-weight:700;color:var(--tx2);text-transform:uppercase}
.badge{display:inline-block;padding:2px 8px;border-radius:8px;font-size:10px;font-weight:700}
.b-tratos{background:#d1fae5;color:#065f46}.b-reforma{background:#ede9fe;color:#5b21b6}.b-estoque{background:#dbeafe;color:#1e40af}
/* Form */
.form-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px}
.form-group{display:flex;flex-direction:column;gap:4px}
.form-group label{font-size:11px;font-weight:700;color:var(--tx2);text-transform:uppercase}
.form-group input,.form-group select{border:2px solid var(--brd);border-radius:8px;padding:10px 14px;font-size:14px;font-family:inherit}
.form-group input:focus,.form-group select:focus{border-color:var(--rz);outline:none;box-shadow:0 0 0 3px rgba(0,132,61,.15)}
.btn{border:none;border-radius:8px;padding:12px 24px;font-size:14px;font-weight:700;cursor:pointer;font-family:inherit;transition:all .15s}
.btn-rz{background:var(--rz);color:#fff}.btn-rz:hover{background:var(--rz2)}
.btn-ci{background:#e5e7eb;color:var(--tx)}.btn-ci:hover{background:#d1d5db}
.msg{padding:12px 16px;border-radius:8px;margin-bottom:12px;font-weight:600;font-size:13px;display:none}
.msg.ok{display:block;background:#d1fae5;color:#065f46;border:1px solid #6ee7b7}
.msg.err{display:block;background:#fee2e2;color:#991b1b;border:1px solid #fca5a5}
.nf-list{max-height:400px;overflow-y:auto;margin-top:12px}
.chart-c{position:relative;height:280px}
.rodape{text-align:center;color:var(--tx2);font-size:10px;padding:16px 0;border-top:1px solid var(--brd);margin-top:20px}
@media(max-width:768px){.grid-2{grid-template-columns:1fr}.form-grid{grid-template-columns:1fr}.kpis{grid-template-columns:repeat(2,1fr)}}
""")
    w("</style></head><body>")

    # Topbar
    w('<div class="topbar"><div><h1>Controle de Corretivos</h1><div class="sub">Safra 26/27 — Ipaussu</div></div><div class="sub">Gerado '+gerado+'</div></div>')

    # Tabs
    w('<div class="tabs">')
    w('<div class="tab on" onclick="go(0,this)">📊 Visão Geral</div>')
    w('<div class="tab" onclick="go(1,this)">📋 Controle de Estoque</div>')
    w('<div class="tab" onclick="go(2,this)">🌱 Doses (Tratos/Reforma)</div>')
    w('<div class="tab" onclick="go(3,this)">📝 Lançar NF</div>')
    w('</div>')

    # ═══ PAGE 0: VISÃO GERAL ═══
    w('<div class="pg on" id="pg0">')
    w('<div class="filter-bar"><label>Período NFs:</label><input type="date" id="vDtI" onchange="renderVisao()"><span style="font-size:11px;color:var(--tx2)">até</span><input type="date" id="vDtF" onchange="renderVisao()"><label style="margin-left:8px">Tipo:</label><select id="vFiltTipo" onchange="renderVisao()"><option value="">Todos</option><option value="Tratos">Tratos</option><option value="Reforma">Reforma</option></select><button class="btn btn-ci" style="padding:6px 14px;font-size:11px" onclick="document.getElementById(\'vDtI\').value=\'\';document.getElementById(\'vDtF\').value=\'\';document.getElementById(\'vFiltTipo\').value=\'\';renderVisao()">Limpar</button></div>')
    w('<div class="kpis" id="kpis"></div>')
    w('<div class="grid-2"><div class="card"><div class="card-title">📊 Status Calcário</div><div class="chart-c"><canvas id="chCalc"></canvas></div></div>')
    w('<div class="card"><div class="card-title">📊 Status Gesso</div><div class="chart-c"><canvas id="chGesso"></canvas></div></div></div>')
    w('<div class="grid-2"><div class="card"><div class="card-title">📈 NFs por Mês (Peso)</div><div class="chart-c"><canvas id="chNfMes"></canvas></div></div>')
    w('<div class="card"><div class="card-title">📈 NFs por Insumo</div><div class="chart-c"><canvas id="chNfIns"></canvas></div></div></div>')
    w('<div class="card"><div class="card-title">🏆 Top Fazendas — Maior Saldo Pendente</div><div class="filter-bar" style="margin-bottom:8px"><label>Buscar:</label><input type="text" id="vBuscaTop" placeholder="Fazenda..." oninput="renderTopSaldo()"><label>Insumo:</label><select id="vFiltIns" onchange="renderTopSaldo()"><option value="calc">Calcário</option><option value="gesso">Gesso</option></select><label>Status:</label><select id="vFiltSt" onchange="renderTopSaldo()"><option value="">Todos</option><option value="bad">✗ Não apontado</option><option value="ok">✓ Apontado</option></select></div><div style="overflow-x:auto;max-height:450px"><table id="tTopSaldo"><tr><td>Carregando...</td></tr></table></div></div>')
    w('</div>')

    # ═══ PAGE 1: CONTROLE DE ESTOQUE ═══
    w('<div class="pg" id="pg1">')
    w('<div class="filter-bar"><label>Buscar:</label><input type="text" id="buscaEst" placeholder="Fazenda, zona, código..." oninput="renderEstoque()"><label>Tipo:</label><select id="filtTipo" onchange="renderEstoque()"><option value="">Todos</option><option value="Tratos">Tratos</option><option value="Reforma">Reforma</option><option value="Estoque">Estoque</option></select><label>Status Calc:</label><select id="filtStCalc" onchange="renderEstoque()"><option value="">Todos</option><option value="ok">✓ Apontado</option><option value="bad">✗ Não apontado</option><option value="est">Estoque</option></select></div>')
    w('<div class="card" style="padding:0;overflow:auto;max-height:75vh"><table id="tEst"><tr><td>Carregando...</td></tr></table></div>')
    w('</div>')

    # ═══ PAGE 2: DOSES ═══
    w('<div class="pg" id="pg2">')
    w('<div class="filter-bar"><label>Buscar:</label><input type="text" id="buscaDose" placeholder="Fazenda, zona, talhão..." oninput="renderDoses()"><label>Fonte:</label><select id="filtFonte" onchange="renderDoses()"><option value="">Todos</option><option value="tratos">Tratos</option><option value="reforma">Reforma</option></select></div>')
    w('<div class="card" style="padding:0;overflow:auto;max-height:75vh"><table id="tDose"><tr><td>Carregando...</td></tr></table></div>')
    w('</div>')

    # ═══ PAGE 3: LANÇAR NF ═══
    w('<div class="pg" id="pg3">')
    w('<div class="card"><div class="card-title">📝 Lançamento de Nota Fiscal</div>')
    w('<div class="msg" id="nfMsg"></div>')
    w('<div class="form-grid">')
    w('<div class="form-group"><label>Número NF</label><input type="number" id="nfNum" placeholder="Ex: 409826"></div>')
    w('<div class="form-group"><label>Cód. Fazenda</label><input type="number" id="nfFaz" placeholder="Ex: 19301" list="fazList"></div>')
    w('<datalist id="fazList">')
    for f in fazendas:
        w(f'<option value="{f["cod"]}">{f["nome"]}</option>')
    w('</datalist>')
    w('<div class="form-group"><label>Zona</label><input type="number" id="nfZona" placeholder="Ex: 283"></div>')
    w('<div class="form-group"><label>Insumo</label><select id="nfInsumo"><option value="Calcario">Calcário</option><option value="Gesso">Gesso</option></select></div>')
    w('<div class="form-group"><label>Peso (Toneladas)</label><input type="number" step="0.01" id="nfPeso" placeholder="Ex: 33.13"></div>')
    w('</div>')
    w('<div style="margin-top:16px;display:flex;gap:10px">')
    w('<button class="btn btn-rz" onclick="salvarNF()">✅ Salvar NF</button>')
    w('<button class="btn btn-ci" onclick="limparForm()">Limpar</button>')
    w('<button class="btn btn-ci" onclick="exportarCSV()">📥 Exportar CSV</button>')
    w('</div></div>')
    w('<div class="card"><div class="card-title">📋 NFs Lançadas nesta Sessão</div><div class="nf-list"><table id="tNfSessao"><thead><tr><th>NF</th><th>Fazenda</th><th>Zona</th><th>Insumo</th><th>Peso (T)</th><th>Data</th></tr></thead><tbody id="nfBody"></tbody></table></div></div>')
    w('<div class="card"><div class="card-title">📋 Histórico de NFs (Base Notas)</div><div style="overflow:auto;max-height:400px"><table id="tNfHist"><tr><td>Carregando...</td></tr></table></div></div>')
    w('</div>')

    # ═══ DATA ═══
    w(f'<script type="application/json" id="d-est">{inj(estoque)}</script>')
    w(f'<script type="application/json" id="d-notas">{inj(notas)}</script>')
    w(f'<script type="application/json" id="d-tratos">{inj(tratos)}</script>')
    w(f'<script type="application/json" id="d-reforma">{inj(reforma)}</script>')
    w(f'<script type="application/json" id="d-faz">{inj(fazendas)}</script>')

    w("<script>")
    w("if(typeof ChartDataLabels!=='undefined')Chart.register(ChartDataLabels);")
    w("""
var EST=[],NOTAS=[],TRATOS=[],REFORMA=[],FAZ=[];
var sessaoNFs=[];
var _charts={};
function mkChart(id,cfg){if(_charts[id])_charts[id].destroy();_charts[id]=new Chart(document.getElementById(id),cfg);return _charts[id]}

function go(i,el){document.querySelectorAll('.pg').forEach((p,j)=>p.classList.toggle('on',j===i));document.querySelectorAll('.tab').forEach(t=>t.classList.remove('on'));el.classList.add('on');if(i===0)renderVisao();if(i===1)renderEstoque();if(i===2)renderDoses();if(i===3)renderNfHist()}

function fmt(n,d){if(isNaN(n)||n===null)return'0';d=d===undefined?2:d;var s=Math.abs(n).toFixed(d),p=s.split('.'),i=p[0],r='',c=0;for(var j=i.length-1;j>=0;j--){if(c>0&&c%3===0)r='.'+r;r=i[j]+r;c++}return(n<0?'-':'')+r+(d>0?','+p[1]:'')}

function stBadge(st){
  if(!st)return'<span class="st-na">—</span>';
  if(st.indexOf('✓')>-1||st.indexOf('Apontado')>-1)return'<span class="st-ok">✓ Apontado</span>';
  if(st.indexOf('✗')>-1||st.indexOf('Não')>-1)return'<span class="st-bad">✗ Não apontado</span>';
  if(st.indexOf('ESTOQUE')>-1)return st.indexOf('SEM')>-1?'<span class="st-bad">Sem Estoque</span>':'<span class="st-stock">Tem Estoque</span>';
  return'<span class="st-na">'+st+'</span>';
}
function tipoBadge(t){
  if(t==='Tratos')return'<span class="badge b-tratos">Tratos</span>';
  if(t==='Reforma')return'<span class="badge b-reforma">Reforma</span>';
  return'<span class="badge b-estoque">'+t+'</span>';
}

// ═══ VISÃO GERAL ═══
function renderVisao(){
  var vTipo=document.getElementById('vFiltTipo')?document.getElementById('vFiltTipo').value:'';
  var tRecC=0,tRecG=0,tNfC=0,tNfG=0,tApC=0,tApG=0,nOkC=0,nBadC=0,nOkG=0,nBadG=0,nT=0,nR=0,nE=0;
  EST.forEach(e=>{
    if(vTipo&&e.tipo!==vTipo)return;
    if(e.tipo==='Tratos')nT++;else if(e.tipo==='Reforma')nR++;else nE++;
    tRecC+=e.recCalcT+(e.recCalcR/1000);tRecG+=(e.recGessoR/1000)+e.recGessoT;
    tNfC+=e.nfCalc;tNfG+=e.nfGesso;tApC+=e.apCalc;tApG+=e.apGesso;
    if(e.stCalc.indexOf('✓')>-1)nOkC++;else if(e.stCalc.indexOf('✗')>-1)nBadC++;
    if(e.stGesso.indexOf('✓')>-1)nOkG++;else if(e.stGesso.indexOf('✗')>-1)nBadG++;
  });
  var h='';
  h+='<div class="kpi"><div class="kpi-lb">Rec. Calcário</div><div class="kpi-vl">'+fmt(tRecC,0)+' T</div><div class="kpi-sub">Tratos + Reforma</div></div>';
  h+='<div class="kpi bl"><div class="kpi-lb">NF Recebidas (Calc)</div><div class="kpi-vl">'+fmt(tNfC,0)+' T</div></div>';
  h+='<div class="kpi ok"><div class="kpi-lb">Apontado (Calc)</div><div class="kpi-vl">'+fmt(tApC,0)+' T</div></div>';
  h+='<div class="kpi am"><div class="kpi-lb">Saldo Calc (Rec-Ap)</div><div class="kpi-vl">'+fmt(tRecC-tApC,0)+' T</div></div>';
  h+='<div class="kpi"><div class="kpi-lb">Rec. Gesso</div><div class="kpi-vl">'+fmt(tRecG,0)+' T</div></div>';
  h+='<div class="kpi bl"><div class="kpi-lb">NF Recebidas (Gesso)</div><div class="kpi-vl">'+fmt(tNfG,0)+' T</div></div>';
  h+='<div class="kpi red"><div class="kpi-lb">Fazendas</div><div class="kpi-vl">'+(nT+nR+nE)+'</div><div class="kpi-sub">'+nT+' tratos · '+nR+' reforma · '+nE+' estoque</div></div>';
  h+='<div class="kpi ok"><div class="kpi-lb">Calc ✓ Apontado</div><div class="kpi-vl">'+nOkC+' / '+(nOkC+nBadC)+'</div><div class="kpi-sub">'+fmt(nOkC/(nOkC+nBadC||1)*100,1)+'%</div></div>';
  document.getElementById('kpis').innerHTML=h;

  // Filter NFs by date
  var di=document.getElementById('vDtI')?document.getElementById('vDtI').value:'';
  var df=document.getElementById('vDtF')?document.getElementById('vDtF').value:'';
  var nfFilt=NOTAS.filter(n=>{
    if(di&&n.data<di)return false;
    if(df&&n.data>df)return false;
    return true;
  });

  // Status charts
  mkChart('chCalc',{type:'doughnut',data:{labels:['✓ Apontado','✗ Não apontado'],datasets:[{data:[nOkC,nBadC],backgroundColor:['#10B981','#EF4444'],borderWidth:0}]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'bottom'},datalabels:{color:'#fff',font:{weight:'bold',size:14},formatter:(v,c)=>v>0?v:''}}}});
  mkChart('chGesso',{type:'doughnut',data:{labels:['✓ Apontado','✗ Não apontado'],datasets:[{data:[nOkG,nBadG],backgroundColor:['#10B981','#EF4444'],borderWidth:0}]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'bottom'},datalabels:{color:'#fff',font:{weight:'bold',size:14},formatter:(v,c)=>v>0?v:''}}}});

  // NFs por mês (filtered)
  var meses={};nfFilt.forEach(n=>{var m=n.mes||'?';if(!meses[m])meses[m]={c:0,g:0};if(n.insumo==='Calcario')meses[m].c+=n.peso;else meses[m].g+=n.peso});
  var mk=Object.keys(meses);
  mkChart('chNfMes',{type:'bar',data:{labels:mk,datasets:[{label:'Calcário (T)',data:mk.map(m=>meses[m].c),backgroundColor:'#00843D'},{label:'Gesso (T)',data:mk.map(m=>meses[m].g),backgroundColor:'#3B82F6'}]},options:{responsive:true,maintainAspectRatio:false,scales:{x:{stacked:true},y:{stacked:true}},plugins:{datalabels:{display:false}}}});

  // NFs por insumo (filtered)
  var ins={};nfFilt.forEach(n=>{var k=n.insumo||'?';if(!ins[k])ins[k]=0;ins[k]+=n.peso});
  var ik=Object.keys(ins);
  mkChart('chNfIns',{type:'pie',data:{labels:ik,datasets:[{data:ik.map(k=>ins[k]),backgroundColor:['#00843D','#3B82F6','#F59E0B','#8B5CF6']}]},options:{responsive:true,maintainAspectRatio:false,plugins:{datalabels:{color:'#fff',font:{weight:'bold'},formatter:(v)=>fmt(v,1)+' T'}}}});

  renderTopSaldo();
}

function renderTopSaldo(){
  var q=(document.getElementById('vBuscaTop')?document.getElementById('vBuscaTop').value:'').toLowerCase();
  var insF=document.getElementById('vFiltIns')?document.getElementById('vFiltIns').value:'calc';
  var stF=document.getElementById('vFiltSt')?document.getElementById('vFiltSt').value:'';
  var isCalc=insF==='calc';
  var filtered=EST.filter(e=>{
    if(e.tipo==='Estoque')return false;
    if(q&&(e.nome+' '+e.cod).toLowerCase().indexOf(q)<0)return false;
    var st=isCalc?e.stCalc:e.stGesso;
    if(stF==='ok'&&st.indexOf('✓')<0)return false;
    if(stF==='bad'&&st.indexOf('✗')<0)return false;
    return true;
  }).sort((a,b)=>{
    var sa=isCalc?b.saldoCalc:b.saldoGesso;
    var sb=isCalc?a.saldoCalc:a.saldoGesso;
    return sa-sb;
  }).slice(0,25);
  var lbl=isCalc?'Calcário':'Gesso';
  var th='<thead><tr><th>Fazenda</th><th>Zona</th><th>Tipo</th><th class="num">Rec. '+lbl+' (T)</th><th class="num">Apontado (T)</th><th class="num">Saldo (T)</th><th>Status</th></tr></thead><tbody>';
  filtered.forEach(e=>{
    var rec=isCalc?e.recCalcT+(e.recCalcR/1000):(e.recGessoR/1000)+e.recGessoT;
    var ap=isCalc?e.apCalc:e.apGesso;
    var saldo=isCalc?e.saldoCalc:e.saldoGesso;
    var st=isCalc?e.stCalc:e.stGesso;
    th+='<tr><td><b>'+e.nome+'</b> <span style="color:var(--tx2);font-size:10px">('+e.cod+')</span></td><td>'+e.zona+'</td><td>'+tipoBadge(e.tipo)+'</td><td class="num">'+fmt(rec,1)+'</td><td class="num">'+fmt(ap,1)+'</td><td class="num" style="font-weight:700;color:'+(saldo>0?'var(--am)':'var(--ok)')+'">'+fmt(saldo,1)+'</td><td>'+stBadge(st)+'</td></tr>';
  });
  th+='</tbody>';document.getElementById('tTopSaldo').innerHTML=th;
}

// ═══ CONTROLE ESTOQUE ═══
function renderEstoque(){
  var q=(document.getElementById('buscaEst').value||'').toLowerCase();
  var ft=document.getElementById('filtTipo').value;
  var fsc=document.getElementById('filtStCalc').value;
  var rows=EST.filter(e=>{
    if(q&&(e.nome+' '+e.cod+' '+e.zona).toLowerCase().indexOf(q)<0)return false;
    if(ft&&e.tipo!==ft)return false;
    if(fsc==='ok'&&e.stCalc.indexOf('✓')<0)return false;
    if(fsc==='bad'&&e.stCalc.indexOf('✗')<0)return false;
    if(fsc==='est'&&e.stCalc.indexOf('ESTOQUE')<0)return false;
    return true;
  });
  var h='<thead><tr><th>Cód.</th><th>Fazenda</th><th>Zona</th><th>Tipo</th><th class="num">Rec.Calc (T)</th><th class="num">Rec.Calc Ref (kg)</th><th class="num">NF Calc (T)</th><th class="num">Ap. Calc (T)</th><th class="num">Saldo Calc</th><th>St. Calc</th><th class="num">Rec.Gesso (T)</th><th class="num">NF Gesso (T)</th><th class="num">Ap. Gesso (T)</th><th>St. Gesso</th></tr></thead><tbody>';
  rows.forEach(e=>{
    h+='<tr><td>'+e.cod+'</td><td><b>'+e.nome+'</b></td><td>'+e.zona+'</td><td>'+tipoBadge(e.tipo)+'</td>';
    h+='<td class="num">'+fmt(e.recCalcT,1)+'</td><td class="num">'+fmt(e.recCalcR,0)+'</td>';
    h+='<td class="num">'+fmt(e.nfCalc,1)+'</td><td class="num">'+fmt(e.apCalc,1)+'</td>';
    h+='<td class="num" style="font-weight:700;color:'+(e.saldoCalc>0?'var(--am)':'var(--ok)')+'">'+fmt(e.saldoCalc,1)+'</td>';
    h+='<td>'+stBadge(e.stCalc)+'</td>';
    h+='<td class="num">'+fmt(e.recGessoT+e.recGessoR/1000,1)+'</td><td class="num">'+fmt(e.nfGesso,1)+'</td><td class="num">'+fmt(e.apGesso,1)+'</td>';
    h+='<td>'+stBadge(e.stGesso)+'</td></tr>';
  });
  h+='</tbody>';document.getElementById('tEst').innerHTML=h;
}

// ═══ DOSES ═══
function renderDoses(){
  var q=(document.getElementById('buscaDose').value||'').toLowerCase();
  var ff=document.getElementById('filtFonte').value;
  var h='<thead><tr><th>Fonte</th><th>Fazenda</th><th>Zona</th><th>Talhão</th><th class="num">Área (ha)</th><th class="num">Dose Calc (kg/ha)</th><th class="num">Vol. Calc (T)</th><th class="num">Taxa Gesso (kg/ha)</th><th class="num">Vol. Gesso (kg)</th><th>Info</th></tr></thead><tbody>';
  if(ff!=='reforma'){
    TRATOS.filter(t=>{if(q&&(t.fazenda+' '+t.cod+' '+t.zona+' '+t.talhao).toLowerCase().indexOf(q)<0)return false;return t.progCalc==='Sim'}).forEach(t=>{
      h+='<tr><td><span class="badge b-tratos">Tratos</span></td><td><b>'+t.fazenda+'</b></td><td>'+t.zona+'</td><td>'+t.talhao+'</td><td class="num">'+fmt(t.areaCalc,2)+'</td><td class="num">'+fmt(t.dose,0)+'</td><td class="num" style="font-weight:700">'+fmt(t.volume,2)+'</td><td class="num">—</td><td class="num">—</td><td style="font-size:10px;color:var(--tx2)">'+t.varied+' · '+t.estagio+'</td></tr>';
    });}
  if(ff!=='tratos'){
    REFORMA.filter(r=>{if(q&&(r.zona+' '+r.talhao+' '+r.tipo).toLowerCase().indexOf(q)<0)return false;return true}).forEach(r=>{
      h+='<tr><td><span class="badge b-reforma">Reforma</span></td><td>—</td><td>'+r.zona+'</td><td>'+r.talhao+'</td><td class="num">'+fmt(r.area,2)+'</td><td class="num">'+fmt(r.txCalc,0)+'</td><td class="num" style="font-weight:700">'+fmt(r.qtdCalc/1000,2)+'</td><td class="num">'+fmt(r.txGesso,0)+'</td><td class="num">'+fmt(r.qtdGesso/1000,2)+'</td><td style="font-size:10px;color:var(--tx2)">'+r.tipo+'</td></tr>';
    });}
  h+='</tbody>';document.getElementById('tDose').innerHTML=h;
}

// ═══ NF LANÇAMENTO ═══
function salvarNF(){
  var nf=document.getElementById('nfNum').value;
  var faz=document.getElementById('nfFaz').value;
  var zona=document.getElementById('nfZona').value;
  var ins=document.getElementById('nfInsumo').value;
  var peso=document.getElementById('nfPeso').value;
  var msg=document.getElementById('nfMsg');
  if(!nf||!faz||!zona||!peso){msg.className='msg err';msg.textContent='Preencha todos os campos.';return}
  var entry={nf:nf,fazCod:faz,zona:zona,insumo:ins,peso:parseFloat(peso),data:new Date().toISOString().slice(0,10)};
  // Find fazenda name
  var fn=FAZ.find(f=>f.cod===faz);
  entry.fazNome=fn?fn.nome:'Cód '+faz;
  entry.chave=faz+zona;
  sessaoNFs.push(entry);
  renderNfSessao();
  msg.className='msg ok';msg.textContent='✅ NF '+nf+' adicionada! Use "Exportar CSV" para importar na planilha.';
  // Try to send to server
  fetch('/api/nf',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(entry)}).then(r=>r.json()).then(d=>{if(d.ok)msg.textContent='✅ NF '+nf+' salva na planilha com sucesso!'}).catch(()=>{});
}
function limparForm(){['nfNum','nfFaz','nfZona','nfPeso'].forEach(id=>document.getElementById(id).value='');document.getElementById('nfMsg').className='msg'}
function renderNfSessao(){
  var h='';sessaoNFs.forEach(e=>{h+='<tr><td>'+e.nf+'</td><td>'+e.fazNome+' ('+e.fazCod+')</td><td>'+e.zona+'</td><td>'+e.insumo+'</td><td class="num">'+fmt(e.peso,2)+'</td><td>'+e.data+'</td></tr>'});
  document.getElementById('nfBody').innerHTML=h;
}
function exportarCSV(){
  if(!sessaoNFs.length){alert('Nenhuma NF para exportar.');return}
  var csv='Chavessig;NF/saida;Data;Fazenda;Zona;Insumo;Peso\\n';
  sessaoNFs.forEach(e=>{csv+=e.chave+';'+e.nf+';'+e.data+';'+e.fazCod+';'+e.zona+';'+e.insumo+';'+e.peso+'\\n'});
  var blob=new Blob([csv],{type:'text/csv'});var a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download='nfs_lancadas_'+new Date().toISOString().slice(0,10)+'.csv';a.click();
}
function renderNfHist(){
  var h='<thead><tr><th>NF</th><th>Data</th><th>Fazenda</th><th>Zona</th><th>Insumo</th><th class="num">Peso (T)</th><th>Mês</th></tr></thead><tbody>';
  NOTAS.slice().reverse().forEach(n=>{
    var dt=n.data?n.data.split('-').reverse().join('/'):'';
    h+='<tr><td><b>'+n.nf+'</b></td><td>'+dt+'</td><td>'+n.fazNome+' ('+n.fazCod+')</td><td>'+n.zona+'</td><td>'+n.insumo+'</td><td class="num">'+fmt(n.peso,2)+'</td><td>'+n.mes+'</td></tr>';
  });
  h+='</tbody>';document.getElementById('tNfHist').innerHTML=h;
}

window.addEventListener('DOMContentLoaded',function(){
  try{
    EST=JSON.parse(document.getElementById('d-est').textContent||'[]');
    NOTAS=JSON.parse(document.getElementById('d-notas').textContent||'[]');
    TRATOS=JSON.parse(document.getElementById('d-tratos').textContent||'[]');
    REFORMA=JSON.parse(document.getElementById('d-reforma').textContent||'[]');
    FAZ=JSON.parse(document.getElementById('d-faz').textContent||'[]');
    renderVisao();
  }catch(e){console.error(e)}
});
""")
    w("</script>")
    w('<div class="rodape">Controle de Corretivos SF 26/27 — Gerado em '+gerado+'</div>')
    w("</body></html>")

    html = "\n".join(L)

    try:
        from auth_dashboard import proteger_html, deve_proteger
        if deve_proteger(): html = proteger_html(html)
    except ImportError: pass

    Path(ARQUIVO_HTML).write_text(html, encoding="utf-8")
    tam = round(Path(ARQUIVO_HTML).stat().st_size / 1024, 1)
    print(f"  [OK] {ARQUIVO_HTML} ({tam} KB)")
    return html


# ═══ SERVIDOR LOCAL (modo --server) ═══
def iniciar_servidor(port=5500):
    from http.server import HTTPServer, SimpleHTTPRequestHandler
    import urllib.parse

    print(f"\n  🚀 Servidor iniciado em http://localhost:{port}")
    print(f"  📝 Lançamentos de NF serão salvos em: {ARQUIVO_EXCEL}")
    print(f"  Ctrl+C para parar\n")

    # Generate fresh HTML
    estoque, notas, tratos, reforma, fazendas = ler_dados()
    gerar_html(estoque, notas, tratos, reforma, fazendas)

    class Handler(SimpleHTTPRequestHandler):
        def do_POST(self):
            if self.path == '/api/nf':
                length = int(self.headers['Content-Length'])
                body = json.loads(self.rfile.read(length))
                try:
                    wb = load_workbook(ARQUIVO_EXCEL)
                    ws = wb['Base Notas ']
                    nr = ws.max_row + 1
                    ws.cell(nr, 1, str(body.get('chave','')))
                    ws.cell(nr, 2, int(body['nf']))
                    ws.cell(nr, 4, int(body['fazCod']))
                    ws.cell(nr, 5, body.get('fazNome',''))
                    ws.cell(nr, 6, int(body['zona']))
                    ws.cell(nr, 7, body['insumo'])
                    ws.cell(nr, 8, float(body['peso']))
                    ws.cell(nr, 9, datetime.now())
                    ws.cell(nr, 10, datetime.now().strftime('%B').lower())
                    wb.save(ARQUIVO_EXCEL)
                    self.send_response(200)
                    self.send_header('Content-Type','application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({'ok':True,'row':nr}).encode())
                    print(f"  ✅ NF {body['nf']} salva na linha {nr}")
                except Exception as e:
                    self.send_response(500)
                    self.send_header('Content-Type','application/json')
                    self.end_headers()
                    self.wfile.write(json.dumps({'ok':False,'error':str(e)}).encode())
                    print(f"  ❌ Erro: {e}")
            else:
                self.send_response(404)
                self.end_headers()

        def do_GET(self):
            if self.path == '/' or self.path == '/index.html':
                self.path = '/' + ARQUIVO_HTML
            super().do_GET()

        def log_message(self, format, *args):
            pass  # Silencia logs HTTP

    HTTPServer(('0.0.0.0', port), Handler).serve_forever()


if __name__ == "__main__":
    import argparse

    try:
        from auth_dashboard import gerenciar_usuarios
        if gerenciar_usuarios(): exit()
    except ImportError: pass

    parser = argparse.ArgumentParser(description='Dashboard Corretivos SF 26/27')
    parser.add_argument('--server', action='store_true', help='Iniciar servidor com API de NF')
    parser.add_argument('-p', '--port', type=int, default=5500, help='Porta do servidor (default: 5500)')
    args, _ = parser.parse_known_args()

    pasta = os.path.dirname(os.path.abspath(__file__))
    os.chdir(pasta)

    if not os.path.exists(ARQUIVO_EXCEL):
        print(f"  [ERRO] {ARQUIVO_EXCEL} não encontrado!")
        sys.exit(1)

    if args.server:
        iniciar_servidor(args.port)
    else:
        estoque, notas, tratos, reforma, fazendas = ler_dados()
        gerar_html(estoque, notas, tratos, reforma, fazendas)

        # Push to GitHub
        os.system(f'git add "{ARQUIVO_HTML}"')
        os.system(f'git commit -m "Corretivo: {time.strftime("%d/%m/%Y %H:%M")}"')
        os.system('git push')

        # Monitor
        if not os.path.exists(ARQUIVO_EXCEL): sys.exit(0)
        ultima_mod = os.path.getmtime(ARQUIVO_EXCEL)
        print(f"  Monitorando {ARQUIVO_EXCEL}...")
        try:
            while True:
                time.sleep(5)
                mod = os.path.getmtime(ARQUIVO_EXCEL)
                if mod != ultima_mod:
                    ultima_mod = mod
                    print(f"\n  [{time.strftime('%H:%M:%S')}] Excel atualizado! Regenerando...")
                    time.sleep(3)
                    try:
                        estoque, notas, tratos, reforma, fazendas = ler_dados()
                        gerar_html(estoque, notas, tratos, reforma, fazendas)
                        os.system(f'git add "{ARQUIVO_HTML}"')
                        os.system(f'git commit -m "Corretivo: {time.strftime("%d/%m/%Y %H:%M")}"')
                        os.system('git push')
                    except Exception as e:
                        print(f"  [ERRO] {e}")
                else:
                    print(f"\r  Monitorando... {time.strftime('%H:%M:%S')}", end="", flush=True)
        except KeyboardInterrupt:
            print("\n  Monitor encerrado.")
